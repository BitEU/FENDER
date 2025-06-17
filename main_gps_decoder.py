import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import sys
import json
import csv
from pathlib import Path
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl import Workbook
import platform
import argparse
from typing import List, Dict, Type
import importlib
import inspect
from datetime import datetime
import hashlib
import shutil
import locale
import socket
import subprocess
import tempfile
import secrets
import stat
import logging
from logging.handlers import RotatingFileHandler
import traceback

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

# FENDER Version Information
FENDER_VERSION = "0.2.2"
FENDER_BUILD_DATE = "June 17 2025"

# Maximum file size (in GB) the program will load (I see Hondas top out at 30, Hyundais in the 180s)
sizeingb = 200

# Duplicate filtering precision
decimals_of_prec = 4

# Setup comprehensive logging
def setup_logging():
    """Setup comprehensive logging with custom timestamp format that appends to a single log file"""
    # Create logs directory if it doesn't exist
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    # Custom formatter with exact timestamp format requested
    class CustomFormatter(logging.Formatter):
        def formatTime(self, record, datefmt=None):
            dt = datetime.fromtimestamp(record.created)
            return dt.strftime('[%Y-%B-%d %H:%M:%S]')
        
        def format(self, record):
            # Get the custom timestamp
            record.asctime = self.formatTime(record)
            # Format the message
            return f"{record.asctime} [{record.levelname}] {record.name} - {record.getMessage()}"
    
    # Configure root logger
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    
    # Remove existing handlers
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # File handler with append mode (no rotation)
    file_handler = logging.FileHandler(
        log_dir / 'fender.log',
        mode='a',  # 'a' for append mode
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(CustomFormatter())
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(CustomFormatter())
    
    # Add handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # Log startup
    logger.info("="*80)
    logger.info(f"FENDER v{FENDER_VERSION} - Forensic Extraction of Navigational Data & Event Records")
    logger.info(f"Build Date: {FENDER_BUILD_DATE}")
    logger.info(f"Python Version: {sys.version}")
    logger.info(f"Platform: {platform.platform()}")
    logger.info(f"Process ID: {os.getpid()}")
    logger.info("="*80)
    
    return logger

# Initialize logging
logger = setup_logging()

# Import base decoder
from base_decoder import BaseDecoder, GPSEntry

if platform.system() == "Windows":
    import ctypes
    from ctypes import windll
    logger.debug("Windows platform detected, imported ctypes and windll")

def get_system_info(input_file=None, output_file=None, execution_mode="GUI", decoder_registry=None):
    """Gather system and configuration information for reports"""
    logger.info("Gathering system information for report generation")
    logger.debug(f"Input file: {input_file}, Output file: {output_file}, Mode: {execution_mode}")
    
    # Get directory paths for permission checking
    output_dir = os.path.dirname(output_file) if output_file else os.getcwd()
    logger.debug(f"Output directory: {output_dir}")
    
    system_info = {
        "fender_version": FENDER_VERSION,
        "fender_build_date": FENDER_BUILD_DATE,
        "report_generated_on": datetime.now().isoformat(),
        "python_interpreter_version": sys.version,
        "python_interpreter_path": sys.executable,
        "operating_system": platform.system(),
        "os_release": platform.release(),
        "os_version": platform.version(),
        "system_architecture": platform.machine(),
        "processor_type": platform.processor(),
        "computer_hostname": platform.node(),
        "system_ram_available_total": get_system_ram(),
        "output_disk_space_available": get_disk_space(output_dir),
        "system_locale": get_system_locale(),
        "network_status": check_network_status(),
        "execution_mode": execution_mode,
    }
    
    logger.debug(f"Basic system info collected: OS={system_info['operating_system']}, "
                f"Architecture={system_info['system_architecture']}")
    
    # Add file permission checks if files are provided
    if input_file:
        logger.debug(f"Checking read permissions for: {input_file}")
        system_info["read_permissions_granted"] = check_read_permissions(input_file)
    
    if output_file:
        logger.debug(f"Checking write permissions for: {output_dir}")
        system_info["write_permissions_granted"] = check_write_permissions(output_dir)
    
    # Add CLI arguments if running in CLI mode
    if execution_mode == "CLI":
        cli_args = " ".join(sys.argv)
        logger.debug(f"CLI arguments: {cli_args}")
        system_info["cli_arguments"] = cli_args
    
    # Add decoder information if registry is provided
    if decoder_registry:
        logger.debug("Collecting decoder information from registry")
        system_info["available_decoders"] = list(decoder_registry.get_decoder_names())
        system_info["decoder_details"] = get_decoder_info(decoder_registry)
        system_info["decoder_hashes"] = get_decoder_hashes(decoder_registry)
        logger.info(f"Found {len(system_info['available_decoders'])} decoders")
    
    # Add file hashes for main components
    try:
        main_script_path = os.path.abspath(__file__)
        logger.debug(f"Calculating hash for main script: {main_script_path}")
        system_info["main_script_hash"] = get_file_hash_safe(main_script_path)
        system_info["main_script_path"] = main_script_path
    except Exception as e:
        logger.error(f"Error getting main script hash: {e}")
        system_info["main_script_hash"] = "Error getting main script hash"
    
    try:
        base_decoder_path = os.path.join(os.path.dirname(__file__), "base_decoder.py")
        if os.path.exists(base_decoder_path):
            logger.debug(f"Calculating hash for base decoder: {base_decoder_path}")
            system_info["base_decoder_hash"] = get_file_hash_safe(base_decoder_path)
            system_info["base_decoder_path"] = base_decoder_path
        else:
            logger.warning(f"base_decoder.py not found at: {base_decoder_path}")
            system_info["base_decoder_hash"] = "base_decoder.py not found"
    except Exception as e:
        logger.error(f"Error getting base decoder hash: {e}")
        system_info["base_decoder_hash"] = "Error getting base decoder hash"
    
    logger.info("System information gathering completed successfully")
    return system_info

def get_decoder_hashes(registry):
    """Get SHA256 hashes of all loaded decoder files for integrity verification"""
    logger.info("Calculating hashes for decoder integrity verification")
    decoder_hashes = {}
    
    for name in registry.get_decoder_names():
        logger.debug(f"Processing decoder: {name}")
        try:
            decoder_class = registry.get_decoder(name)
            
            # Get the module file path
            module = inspect.getmodule(decoder_class)
            if module and hasattr(module, '__file__') and module.__file__:
                file_path = os.path.abspath(module.__file__)
                logger.debug(f"Decoder {name} located at: {file_path}")
                
                # Calculate hash
                decoder_hashes[name] = {
                    "file_path": file_path,
                    "sha256_hash": get_file_hash_safe(file_path),
                    "file_size": os.path.getsize(file_path) if os.path.exists(file_path) else 0,
                    "last_modified": datetime.fromtimestamp(
                        os.path.getmtime(file_path)
                    ).isoformat() if os.path.exists(file_path) else "Unknown"
                }
                logger.debug(f"Hash for {name}: {decoder_hashes[name]['sha256_hash'][:16]}...")
            else:
                logger.warning(f"Could not determine file path for decoder: {name}")
                decoder_hashes[name] = {
                    "error": "Could not determine decoder file path"
                }
                
        except Exception as e:
            logger.error(f"Error getting decoder hash for {name}: {e}", exc_info=True)
            decoder_hashes[name] = {
                "error": f"Error getting decoder hash: {str(e)}"
            }
    
    logger.info(f"Completed hash calculation for {len(decoder_hashes)} decoders")
    return decoder_hashes

def secure_temp_file(suffix="", prefix="fender_", dir=None):
    """Create a secure temporary file with restricted permissions"""
    logger.debug(f"Creating secure temporary file with prefix={prefix}, suffix={suffix}")
    
    # Create temporary file with secure permissions
    fd, path = tempfile.mkstemp(suffix=suffix, prefix=prefix, dir=dir)
    logger.debug(f"Created temporary file: {path}")
    
    # Set restrictive permissions (owner read/write only)
    if platform.system() != "Windows":
        logger.debug(f"Setting restrictive permissions on temporary file")
        os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)
    
    logger.info(f"Secure temporary file created: {path}")
    return fd, path

def secure_temp_dir(prefix="fender_", dir=None):
    """Create a secure temporary directory with restricted permissions"""
    logger.debug(f"Creating secure temporary directory with prefix={prefix}")
    
    path = tempfile.mkdtemp(prefix=prefix, dir=dir)
    logger.debug(f"Created temporary directory: {path}")
    
    # Set restrictive permissions (owner only)
    if platform.system() != "Windows":
        logger.debug(f"Setting restrictive permissions on temporary directory")
        os.chmod(path, stat.S_IRWXU)
    
    logger.info(f"Secure temporary directory created: {path}")
    return path

def secure_file_copy(src, dst, chunk_size=65536):
    """Securely copy file with verification"""
    logger.info(f"Starting secure file copy from {src} to {dst}")
    logger.debug(f"Using chunk size: {chunk_size} bytes")
    
    src_hash = hashlib.sha256()
    dst_hash = hashlib.sha256()
    bytes_copied = 0
    
    try:
        with open(src, 'rb') as src_file, open(dst, 'wb') as dst_file:
            while True:
                chunk = src_file.read(chunk_size)
                if not chunk:
                    break
                src_hash.update(chunk)
                dst_file.write(chunk)
                bytes_copied += len(chunk)
                
                if bytes_copied % (chunk_size * 100) == 0:  # Log progress every 100 chunks
                    logger.debug(f"Copied {bytes_copied} bytes...")
        
        logger.debug(f"Total bytes copied: {bytes_copied}")
        
        # Verify copy integrity
        logger.debug("Verifying file copy integrity")
        with open(dst, 'rb') as dst_file:
            while True:
                chunk = dst_file.read(chunk_size)
                if not chunk:
                    break
                dst_hash.update(chunk)
        
        src_hex = src_hash.hexdigest()
        dst_hex = dst_hash.hexdigest()
        
        if src_hex != dst_hex:
            logger.error(f"File copy verification failed! Source hash: {src_hex}, Destination hash: {dst_hex}")
            raise ValueError("File copy verification failed - checksums don't match")
        
        logger.info(f"File copy completed successfully. Hash: {dst_hex}")
        return dst_hex
        
    except Exception as e:
        logger.error(f"Error during secure file copy: {e}", exc_info=True)
        raise

def sanitize_filename(filename):
    """Sanitize filename to prevent path traversal attacks"""
    logger.debug(f"Sanitizing filename: {filename}")
    
    # Remove directory separators and other potentially dangerous characters
    dangerous_chars = '<>:"/\\|?*'
    for char in dangerous_chars:
        filename = filename.replace(char, '_')
    
    # Remove any path components
    filename = os.path.basename(filename)
    
    # Limit length
    if len(filename) > 200:
        name, ext = os.path.splitext(filename)
        filename = name[:200-len(ext)] + ext
        logger.debug(f"Filename truncated to 200 characters")
    
    logger.debug(f"Sanitized filename: {filename}")
    return filename

def validate_file_path(file_path, allowed_extensions=None):
    """Validate file path for security"""
    logger.info(f"Validating file path: {file_path}")
    
    try:
        # Resolve to absolute path to prevent traversal
        abs_path = os.path.abspath(file_path)
        logger.debug(f"Resolved to absolute path: {abs_path}")
        
        # Check if file exists
        if not os.path.exists(abs_path):
            logger.warning(f"File does not exist: {abs_path}")
            return False, "File does not exist"
        
        # Check if it's actually a file
        if not os.path.isfile(abs_path):
            logger.warning(f"Path is not a file: {abs_path}")
            return False, "Path is not a file"
        
        # Check file extension if provided
        if allowed_extensions:
            file_ext = os.path.splitext(abs_path)[1].lower()
            logger.debug(f"File extension: {file_ext}, Allowed: {allowed_extensions}")
            if file_ext not in [ext.lower() for ext in allowed_extensions]:
                logger.warning(f"File extension {file_ext} not in allowed list")
                return False, f"File extension not allowed. Allowed: {allowed_extensions}"
        
        # Check file size (prevent extremely large files)
        file_size = os.path.getsize(abs_path)
        max_size = sizeingb * 1024 * 1024 * 1024
        logger.debug(f"File size: {file_size} bytes, Max allowed: {max_size} bytes")
        
        if file_size > max_size:
            logger.warning(f"File too large: {file_size} bytes")
            return False, f"File too large. Maximum size: {max_size/1024/1024/1024:.1f}GB"
        
        logger.info(f"File validation successful: {abs_path}")
        return True, abs_path
        
    except Exception as e:
        logger.error(f"Path validation error: {e}", exc_info=True)
        return False, f"Path validation error: {str(e)}"

def validate_folder_path(folder_path):
    """Validate folder path for security"""
    logger.info(f"Validating folder path: {folder_path}")
    
    try:
        # Resolve to absolute path to prevent traversal
        abs_path = os.path.abspath(folder_path)
        logger.debug(f"Resolved to absolute path: {abs_path}")
        
        # Check if folder exists
        if not os.path.exists(abs_path):
            logger.warning(f"Folder does not exist: {abs_path}")
            return False, "Folder does not exist"
        
        # Check if it's actually a directory
        if not os.path.isdir(abs_path):
            logger.warning(f"Path is not a folder: {abs_path}")
            return False, "Path is not a folder"
        
        # Check if folder is accessible
        if not os.access(abs_path, os.R_OK):
            logger.warning(f"Folder is not readable: {abs_path}")
            return False, "Folder is not readable"
        
        logger.info(f"Folder validation successful: {abs_path}")
        return True, abs_path
        
    except Exception as e:
        logger.error(f"Folder validation error: {e}", exc_info=True)
        return False, f"Folder validation error: {str(e)}"

def write_geojson(entries: List[GPSEntry], output_path: str, decoder_name: str = "Unknown"):
    """Write GPS entries to GeoJSON format"""
    logger.info(f"Writing {len(entries)} entries to GeoJSON file: {output_path}")
    logger.debug(f"Using decoder: {decoder_name}")
    
    features = []
    skipped_count = 0
    
    for i, entry in enumerate(entries):
        # Skip invalid coordinates
        if (entry.latitude == 0 and entry.longitude == 0) or \
           not (-90 <= entry.latitude <= 90) or \
           not (-180 <= entry.longitude <= 180):
            logger.debug(f"Skipping invalid coordinates at index {i}: lat={entry.latitude}, lon={entry.longitude}")
            skipped_count += 1
            continue
        
        # Create feature
        feature = {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [entry.longitude, entry.latitude]  # GeoJSON uses [lon, lat]
            },
            "properties": {
                "id": i + 1,
                "timestamp": entry.timestamp,
                "latitude": entry.latitude,
                "longitude": entry.longitude,
            }
        }
        
        # Add extra data if available
        if entry.extra_data:
            feature["properties"].update(entry.extra_data)
            logger.debug(f"Added extra data to feature {i+1}: {list(entry.extra_data.keys())}")
        
        features.append(feature)
    
    logger.info(f"Created {len(features)} valid features, skipped {skipped_count} invalid entries")
    
    # Create GeoJSON structure
    geojson = {
        "type": "FeatureCollection",
        "metadata": {
            "decoder": decoder_name,
            "extraction_timestamp": datetime.now().isoformat(),
            "total_features": len(features),
            "coordinate_system": "WGS84",
            "creator": f"FENDER v{FENDER_VERSION}"
        },
        "features": features
    }
    
    # Write to file
    try:
        logger.debug(f"Writing GeoJSON to file")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(geojson, f, indent=2, ensure_ascii=False)
        logger.info(f"GeoJSON file written successfully: {output_path}")
    except Exception as e:
        logger.error(f"Error writing GeoJSON file: {e}", exc_info=True)
        raise

def write_kml(entries: List[GPSEntry], output_path: str, decoder_name: str = "Unknown"):
    """Write GPS entries to KML format for Google Earth"""
    logger.info(f"Writing {len(entries)} entries to KML file: {output_path}")
    logger.debug(f"Using decoder: {decoder_name}")
    
    # KML header with XML declaration
    kml_content = ['<?xml version="1.0" encoding="UTF-8"?>']
    kml_content.append('<kml xmlns="http://www.opengis.net/kml/2.2">')
    kml_content.append('  <Document>')
    
    # Document metadata
    kml_content.append(f'    <name>FENDER GPS Data - {decoder_name}</name>')
    kml_content.append(f'    <description>Extracted by FENDER v{FENDER_VERSION} on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</description>')
    
    # Define styles for placemarks
    kml_content.append('    <Style id="normalPin">')
    kml_content.append('      <IconStyle>')
    kml_content.append('        <color>ff0000ff</color>')  # Red color in KML format (aabbggrr)
    kml_content.append('        <scale>0.8</scale>')
    kml_content.append('        <Icon>')
    kml_content.append('          <href>http://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png</href>')
    kml_content.append('        </Icon>')
    kml_content.append('      </IconStyle>')
    kml_content.append('      <LabelStyle>')
    kml_content.append('        <scale>0.7</scale>')
    kml_content.append('      </LabelStyle>')
    kml_content.append('    </Style>')
    
    # Style for path/track (if we want to connect points)
    kml_content.append('    <Style id="trackStyle">')
    kml_content.append('      <LineStyle>')
    kml_content.append('        <color>ff00ff00</color>')  # Green color
    kml_content.append('        <width>2</width>')
    kml_content.append('      </LineStyle>')
    kml_content.append('    </Style>')
    
    # Create folder for all placemarks
    kml_content.append('    <Folder>')
    kml_content.append('      <name>GPS Locations</name>')
    kml_content.append('      <open>1</open>')
    
    valid_entries = []
    skipped_count = 0
    
    # Add placemarks for each GPS entry
    for i, entry in enumerate(entries):
        # Skip invalid coordinates
        if (entry.latitude == 0 and entry.longitude == 0) or \
           not (-90 <= entry.latitude <= 90) or \
           not (-180 <= entry.longitude <= 180):
            logger.debug(f"Skipping invalid coordinates at index {i}: lat={entry.latitude}, lon={entry.longitude}")
            skipped_count += 1
            continue
        
        valid_entries.append(entry)
        
        kml_content.append('      <Placemark>')
        kml_content.append(f'        <name>Location {i + 1}</name>')
        
        # Build description with all available data
        description_parts = []
        description_parts.append(f'Timestamp: {entry.timestamp}')
        description_parts.append(f'Latitude: {entry.latitude}')
        description_parts.append(f'Longitude: {entry.longitude}')
        
        # Add extra data if available
        if entry.extra_data:
            for key, value in entry.extra_data.items():
                if value and str(value).strip():
                    description_parts.append(f'{key}: {value}')
        
        description = '<![CDATA[' + '<br/>'.join(description_parts) + ']]>'
        kml_content.append(f'        <description>{description}</description>')
        
        kml_content.append('        <styleUrl>#normalPin</styleUrl>')
        
        # Add timestamp if available
        if entry.timestamp and entry.timestamp.strip():
            try:
                # Convert timestamp to ISO format for KML
                # Handle various timestamp formats
                timestamp_str = entry.timestamp.strip()
                if 'T' not in timestamp_str and ' ' in timestamp_str:
                    timestamp_str = timestamp_str.replace(' ', 'T')
                if not timestamp_str.endswith('Z'):
                    timestamp_str += 'Z'
                
                kml_content.append('        <TimeStamp>')
                kml_content.append(f'          <when>{timestamp_str}</when>')
                kml_content.append('        </TimeStamp>')
            except Exception as e:
                logger.debug(f"Could not format timestamp for KML: {e}")
        
        kml_content.append('        <Point>')
        kml_content.append(f'          <coordinates>{entry.longitude},{entry.latitude},0</coordinates>')
        kml_content.append('        </Point>')
        kml_content.append('      </Placemark>')
    
    kml_content.append('    </Folder>')
    
    # Optionally add a path connecting all points (useful for route visualization)
    if len(valid_entries) > 1:
        kml_content.append('    <Placemark>')
        kml_content.append('      <name>GPS Track</name>')
        kml_content.append('      <description>Path connecting all GPS points in chronological order</description>')
        kml_content.append('      <styleUrl>#trackStyle</styleUrl>')
        kml_content.append('      <LineString>')
        kml_content.append('        <tessellate>1</tessellate>')
        kml_content.append('        <coordinates>')
        
        # Sort entries by timestamp for proper path
        sorted_entries = sorted(valid_entries, key=lambda x: x.timestamp if x.timestamp else '')
        
        for entry in sorted_entries:
            kml_content.append(f'          {entry.longitude},{entry.latitude},0')
        
        kml_content.append('        </coordinates>')
        kml_content.append('      </LineString>')
        kml_content.append('    </Placemark>')
    
    kml_content.append('  </Document>')
    kml_content.append('</kml>')
    
    logger.info(f"Created KML with {len(valid_entries)} valid placemarks, skipped {skipped_count} invalid entries")
    
    # Write to file
    try:
        logger.debug(f"Writing KML to file")
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(kml_content))
        logger.info(f"KML file written successfully: {output_path}")
    except Exception as e:
        logger.error(f"Error writing KML file: {e}", exc_info=True)
        raise

def get_file_hash(file_path: str) -> str:
    """Calculate SHA256 hash of the input file or folder"""
    logger.debug(f"Calculating SHA256 hash for: {file_path}")
    
    try:
        sha256_hash = hashlib.sha256()
        
        if os.path.isdir(file_path):
            # For folders, create a hash of all file paths and their modification times
            logger.debug("Hashing folder structure")
            file_list = []
            for root, dirs, files in os.walk(file_path):
                for file in sorted(files):  # Sort for consistent hash
                    file_path_full = os.path.join(root, file)
                    try:
                        mtime = os.path.getmtime(file_path_full)
                        file_list.append(f"{file_path_full}:{mtime}")
                    except:
                        pass
            
            # Hash the file list
            sha256_hash.update("\n".join(file_list).encode('utf-8'))
            hash_result = sha256_hash.hexdigest()
            logger.debug(f"Folder structure hash calculated: {hash_result}")
        else:
            # Original file hashing logic
            bytes_processed = 0
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    sha256_hash.update(chunk)
                    bytes_processed += len(chunk)
            
            hash_result = sha256_hash.hexdigest()
            logger.debug(f"File hash calculated: {hash_result} ({bytes_processed} bytes processed)")
        
        return hash_result
        
    except Exception as e:
        logger.error(f"Error calculating hash for {file_path}: {e}", exc_info=True)
        return f"Error calculating hash: {str(e)}"

def filter_duplicate_entries(entries: List['GPSEntry'], precision: int = 4, logger=None) -> List['GPSEntry']:
    """
    Filter duplicate GPS entries based on timestamp and coordinate precision.
    
    Args:
        entries: List of GPSEntry objects
        precision: Number of decimal places to consider for coordinate comparison
        logger: Logger instance for debug output
        
    Returns:
        Filtered list of GPSEntry objects with duplicates removed
    """
    if not entries:
        return entries
    
    if logger:
        logger.info(f"Starting duplicate filtering with {len(entries)} entries, precision={precision} decimals")
    
    # Group entries by timestamp
    timestamp_groups = {}
    for entry in entries:
        if entry.timestamp not in timestamp_groups:
            timestamp_groups[entry.timestamp] = []
        timestamp_groups[entry.timestamp].append(entry)
    
    if logger:
        logger.debug(f"Grouped entries into {len(timestamp_groups)} unique timestamps")
    
    # Filter duplicates within each timestamp group
    filtered_entries = []
    duplicates_removed = 0
    
    for timestamp, group_entries in timestamp_groups.items():
        if len(group_entries) == 1:
            # Only one entry for this timestamp, keep it
            filtered_entries.append(group_entries[0])
        else:
            # Multiple entries for same timestamp, check coordinates
            kept_entries = []
            
            for entry in group_entries:
                is_duplicate = False
                
                # Check against already kept entries
                for kept_entry in kept_entries:
                    # Round coordinates to specified precision
                    lat1_rounded = round(entry.latitude, precision)
                    lon1_rounded = round(entry.longitude, precision)
                    lat2_rounded = round(kept_entry.latitude, precision)
                    lon2_rounded = round(kept_entry.longitude, precision)
                    
                    # Check if coordinates match at the specified precision
                    if lat1_rounded == lat2_rounded and lon1_rounded == lon2_rounded:
                        is_duplicate = True
                        duplicates_removed += 1
                        break
                
                if not is_duplicate:
                    kept_entries.append(entry)
            
            filtered_entries.extend(kept_entries)
            
            if logger and len(group_entries) > len(kept_entries):
                logger.debug(f"Timestamp {timestamp}: Reduced from {len(group_entries)} to {len(kept_entries)} entries")
    
    # Sort by timestamp to maintain order
    filtered_entries.sort(key=lambda x: x.timestamp)
    
    if logger:
        logger.info(f"Duplicate filtering complete: {len(entries)} -> {len(filtered_entries)} entries ({duplicates_removed} duplicates removed)")
    
    return filtered_entries

def get_extraction_info(decoder_name: str, input_file: str, output_file: str, entry_count: int, processing_time: float = None):
    """Gather extraction-specific information"""
    logger.info("Gathering extraction information")
    logger.debug(f"Decoder: {decoder_name}, Entries: {entry_count}, Time: {processing_time}")
    
    file_size = os.path.getsize(input_file) if os.path.exists(input_file) else 0
    
    extraction_info = {
        "input_file": {
            "path": input_file,
            "filename": os.path.basename(input_file),
            "size_bytes": file_size,
            "size_mb": round(file_size / (1024 * 1024), 2),
            "sha256_hash": get_file_hash(input_file)
        },
        "output_file": {
            "path": output_file,
            "filename": os.path.basename(output_file)
        },
        "extraction_details": {
            "decoder_used": decoder_name,
            "entries_extracted": entry_count,
            "processing_time_seconds": processing_time,
        }
    }
    
    logger.info(f"Extraction info compiled: {entry_count} entries from {extraction_info['input_file']['size_mb']}MB file")
    return extraction_info

def get_system_ram():
    """Get system RAM information"""
    logger.debug("Getting system RAM information")
    
    if PSUTIL_AVAILABLE:
        try:
            memory = psutil.virtual_memory()
            available_gb = memory.available / (1024**3)
            total_gb = memory.total / (1024**3)
            result = f"{available_gb:.1f} GB / {total_gb:.1f} GB"
            logger.debug(f"RAM info via psutil: {result}")
            return result
        except Exception as e:
            logger.error(f"psutil error getting RAM info: {e}")
            return f"psutil error: {str(e)}"
    else:
        logger.debug("psutil not available, using fallback method")
        return get_system_ram_fallback()

def get_system_ram_fallback():
    """Get system RAM information using platform-specific commands"""
    logger.debug("Using fallback method for RAM information")
    
    try:
        if platform.system() == "Windows":
            logger.debug("Using Windows WMI for RAM info")
            result = subprocess.run(['wmic', 'computersystem', 'get', 'TotalPhysicalMemory'], 
                                  capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                lines = result.stdout.strip().split('\n')
                if len(lines) > 1:
                    total_bytes = int(lines[1].strip())
                    total_gb = total_bytes / (1024**3)
                    ram_info = f"Total: {total_gb:.1f} GB (Available: Unknown)"
                    logger.debug(f"Windows RAM info: {ram_info}")
                    return ram_info
                    
        elif platform.system() == "Linux":
            logger.debug("Using /proc/meminfo for RAM info")
            with open('/proc/meminfo', 'r') as f:
                lines = f.readlines()
                mem_total = mem_available = None
                for line in lines:
                    if line.startswith('MemTotal:'):
                        mem_total = int(line.split()[1]) * 1024
                    elif line.startswith('MemAvailable:'):
                        mem_available = int(line.split()[1]) * 1024
                
                if mem_total:
                    total_gb = mem_total / (1024**3)
                    if mem_available:
                        available_gb = mem_available / (1024**3)
                        ram_info = f"{available_gb:.1f} GB / {total_gb:.1f} GB"
                        logger.debug(f"Linux RAM info: {ram_info}")
                        return ram_info
                    else:
                        return f"Total: {total_gb:.1f} GB (Available: Unknown)"
        
        logger.warning("RAM info not available on this platform")
        return "RAM info not available on this platform"
        
    except Exception as e:
        logger.error(f"Error getting RAM info via fallback: {e}", exc_info=True)
        return f"Error getting RAM info: {str(e)}"

def get_disk_space(path):
    """Get available disk space for a given path"""
    logger.debug(f"Getting disk space for path: {path}")
    
    try:
        usage = shutil.disk_usage(os.path.dirname(path))
        available_gb = usage.free / (1024**3)
        result = f"{available_gb:.1f} GB"
        logger.debug(f"Disk space available: {result}")
        return result
    except Exception as e:
        logger.error(f"Error getting disk space: {e}")
        return f"Error getting disk space: {str(e)}"

def check_read_permissions(file_path):
    """Check if file is readable"""
    logger.debug(f"Checking read permissions for: {file_path}")
    
    try:
        result = "Yes" if os.access(file_path, os.R_OK) else "No"
        logger.debug(f"Read permission for {file_path}: {result}")
        return result
    except Exception as e:
        logger.error(f"Error checking read permissions: {e}")
        return "Error checking permissions"

def check_write_permissions(directory_path):
    """Check if directory is writable"""
    logger.debug(f"Checking write permissions for: {directory_path}")
    
    try:
        result = "Yes" if os.access(directory_path, os.W_OK) else "No"
        logger.debug(f"Write permission for {directory_path}: {result}")
        return result
    except Exception as e:
        logger.error(f"Error checking write permissions: {e}")
        return "Error checking permissions"

def get_system_locale():
    """Get system locale information using modern locale methods"""
    logger.debug("Getting system locale information")
    
    try:
        current_locale = locale.getlocale()
        if current_locale[0]:
            locale_info = current_locale[0]
        else:
            try:
                locale.setlocale(locale.LC_ALL, '')
                locale_info = locale.getlocale()[0] or "Unknown"
            except locale.Error:
                try:
                    encoding = locale.getencoding()
                    locale_info = f"Default encoding: {encoding}"
                except:
                    locale_info = "Unknown"
        
        try:
            encoding = locale.getencoding()
            result = f"{locale_info} (Encoding: {encoding})"
        except:
            result = locale_info
        
        logger.debug(f"System locale: {result}")
        return result
            
    except Exception as e:
        logger.error(f"Error getting locale: {e}")
        return f"Error getting locale: {str(e)}"

def check_network_status():
    """Check if system has network connectivity"""
    logger.debug("Checking network connectivity")
    
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=3)
        logger.debug("Network status: Online")
        return "Online"
    except Exception as e:
        logger.debug(f"Network status: Offline - {e}")
        return "Offline"

def get_file_hash_safe(file_path):
    """Get file hash with error handling"""
    logger.debug(f"Safely getting file hash for: {file_path}")
    
    try:
        return get_file_hash(file_path)
    except Exception as e:
        logger.error(f"Error in get_file_hash_safe: {e}")
        return f"Error: {str(e)}"

def get_decoder_info(registry):
    """Get information about available decoders"""
    logger.info("Getting information about available decoders")
    
    decoder_info = {}
    for name in registry.get_decoder_names():
        logger.debug(f"Getting info for decoder: {name}")
        try:
            decoder_class = registry.get_decoder(name)
            decoder_instance = decoder_class()
            decoder_info[name] = {
                "supported_extensions": decoder_instance.get_supported_extensions(),
                "class_name": decoder_class.__name__,
                "module": decoder_class.__module__
            }
            logger.debug(f"Decoder {name}: extensions={decoder_info[name]['supported_extensions']}")
        except Exception as e:
            logger.error(f"Error getting info for decoder {name}: {e}")
            decoder_info[name] = {"error": str(e)}
    
    logger.info(f"Retrieved info for {len(decoder_info)} decoders")
    return decoder_info

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    logger.debug(f"Getting resource path for: {relative_path}")
    
    try:
        base_path = sys._MEIPASS
        logger.debug(f"Running from PyInstaller bundle, base path: {base_path}")
    except Exception:
        base_path = os.path.abspath(".")
        logger.debug(f"Running from source, base path: {base_path}")
        
    resource_path = os.path.join(base_path, relative_path)
    logger.debug(f"Resource path resolved to: {resource_path}")
    return resource_path

class DecoderRegistry:
    """Registry for managing available decoders"""
    def __init__(self):
        logger.info("Initializing DecoderRegistry")
        self.decoders: Dict[str, Type[BaseDecoder]] = {}
        self.auto_discover_decoders()

    def register(self, decoder_class: Type[BaseDecoder]):
        """Register a new decoder"""
        try:
            instance = decoder_class()
            decoder_name = instance.get_name()
            self.decoders[decoder_name] = decoder_class
            logger.info(f"Registered decoder: {decoder_name} ({decoder_class.__name__})")
        except Exception as e:
            logger.error(f"Failed to register decoder {decoder_class}: {e}")

    def auto_discover_decoders(self):
        """Automatically discover and register decoders from the decoders directory"""
        logger.info("Starting auto-discovery of decoders")
        
        if getattr(sys, 'frozen', False):
            decoders_dir = Path(get_resource_path("decoders"))
            resource_path = get_resource_path("")
            if resource_path not in sys.path:
                sys.path.insert(0, resource_path)
                logger.debug(f"Added to sys.path: {resource_path}")
        else:
            decoders_dir = Path("decoders")
            sys.path.append(str(decoders_dir.parent))
            logger.debug(f"Added to sys.path: {decoders_dir.parent}")

        if not decoders_dir.exists():
            logger.error(f"Decoders directory not found: {decoders_dir}")
            return

        logger.info(f"Looking for decoders in: {decoders_dir}")

        decoder_files = list(decoders_dir.glob("*_decoder.py"))
        logger.info(f"Found {len(decoder_files)} potential decoder files")

        for file_path in decoder_files:
            module_name = f"decoders.{file_path.stem}"
            logger.debug(f"Attempting to load module: {module_name}")
            
            try:
                module = importlib.import_module(module_name)
                logger.debug(f"Successfully imported module: {module_name}")
                
                # Look for decoder classes in the module
                for name, obj in inspect.getmembers(module):
                    if (inspect.isclass(obj) and
                        issubclass(obj, BaseDecoder) and
                        obj != BaseDecoder):
                        logger.debug(f"Found decoder class: {name}")
                        self.register(obj)
                        
            except Exception as e:
                logger.error(f"Failed to load decoder from {file_path}: {e}", exc_info=True)

        logger.info(f"Auto-discovery complete. Registered {len(self.decoders)} decoders")

    def get_decoder_names(self) -> List[str]:
        """Get list of available decoder names"""
        names = sorted(self.decoders.keys())
        logger.debug(f"Available decoders: {names}")
        return names

    def get_decoder(self, name: str) -> Type[BaseDecoder]:
        """Get decoder class by name"""
        logger.debug(f"Retrieving decoder: {name}")
        decoder = self.decoders.get(name)
        if not decoder:
            logger.warning(f"Decoder not found: {name}")
        return decoder

class CustomRadiobutton(tk.Canvas):
    """Custom radiobutton that matches the dark theme"""
    def __init__(self, parent, text, variable, value, **kwargs):
        super().__init__(parent, highlightthickness=0, **kwargs)
        logger.debug(f"Creating CustomRadiobutton: text='{text}', value='{value}'")
        
        self.text = text
        self.variable = variable
        self.value = value
        self.selected = False
        
        # Colors
        self.bg_color = '#1a1a1a'
        self.border_color = '#666666'
        self.selected_color = '#4a9eff'
        self.text_color = '#cccccc'
        self.text_selected_color = '#ffffff'
        
        # Calculate width based on text length
        text_width = len(text) * 8 + 40  # Rough estimate: 8px per char + padding
        canvas_width = max(120, text_width)  # Minimum 120px
        
        # Configure canvas
        self.configure(bg=self.bg_color, width=canvas_width, height=30)
        
        # Bind events
        self.bind('<Button-1>', self.on_click)
        self.bind('<Enter>', self.on_enter)
        self.bind('<Leave>', self.on_leave)
        
        # Watch variable changes
        self.variable.trace('w', self.on_variable_change)
        
        # Initial draw
        self.draw()
    
    def draw(self):
        """Draw the radiobutton"""
        self.delete('all')
        
        # Check if selected
        self.selected = (self.variable.get() == self.value)
        
        # Draw circle
        circle_x = 10
        circle_y = 15
        circle_r = 6
        
        if self.selected:
            # Selected state - filled circle
            self.create_oval(circle_x - circle_r, circle_y - circle_r,
                           circle_x + circle_r, circle_y + circle_r,
                           outline=self.selected_color, fill=self.selected_color, width=2)
            # Inner circle
            self.create_oval(circle_x - 3, circle_y - 3,
                           circle_x + 3, circle_y + 3,
                           outline='white', fill='white', width=1)
        else:
            # Unselected state - empty circle
            self.create_oval(circle_x - circle_r, circle_y - circle_r,
                           circle_x + circle_r, circle_y + circle_r,
                           outline=self.border_color, fill=self.bg_color, width=2)
        
        # Draw text
        text_color = self.text_selected_color if self.selected else self.text_color
        self.create_text(25, 15, text=self.text, anchor='w', 
                        fill=text_color, font=('Segoe UI', 10))
    
    def on_click(self, event):
        """Handle click event"""
        logger.debug(f"CustomRadiobutton clicked: {self.text} -> {self.value}")
        self.variable.set(self.value)
    
    def on_enter(self, event):
        """Handle mouse enter"""
        self.configure(cursor='hand2')
    
    def on_leave(self, event):
        """Handle mouse leave"""
        self.configure(cursor='')
    
    def on_variable_change(self, *args):
        """Handle variable change"""
        self.draw()

class CustomToggleButton(tk.Canvas):
    """Custom toggle button that matches the dark theme"""
    def __init__(self, parent, text, variable, **kwargs):
        super().__init__(parent, highlightthickness=0, **kwargs)
        
        self.text = text
        self.variable = variable
        self.is_on = variable.get()
        
        # Colors
        self.bg_color = '#1a1a1a'
        self.off_color = '#666666'
        self.on_color = '#4a9eff'
        self.slider_color = '#ffffff'
        self.text_color = '#cccccc'
        self.text_on_color = '#ffffff'
        
        # Dimensions
        self.toggle_width = 50
        self.toggle_height = 24
        self.slider_size = 18
        
        # Calculate total width
        text_width = len(text) * 8 + 70  # Text + toggle + padding
        canvas_width = max(150, text_width)
        
        # Configure canvas
        self.configure(bg=self.bg_color, width=canvas_width, height=30)
        
        # Bind events
        self.bind('<Button-1>', self.toggle)
        self.bind('<Enter>', self.on_enter)
        self.bind('<Leave>', self.on_leave)
        
        # Watch variable changes
        self.variable.trace('w', self.on_variable_change)
        
        # Initial draw
        self.draw()
    
    def draw(self):
        """Draw the toggle button"""
        self.delete('all')

        # Update state
        self.is_on = self.variable.get()

        # Draw toggle background as a rounded rectangle
        toggle_x = 10
        toggle_y = (30 - self.toggle_height) // 2
        r = self.toggle_height // 2  # radius for rounded ends

        x1 = toggle_x
        y1 = toggle_y
        x2 = toggle_x + self.toggle_width
        y2 = toggle_y + self.toggle_height

        fill_color = self.on_color if self.is_on else self.off_color

        # Draw left arc
        self.create_arc(x1, y1, x1 + 2*r, y2, start=90, extent=180, fill=fill_color, outline=fill_color)
        # Draw right arc
        self.create_arc(x2 - 2*r, y1, x2, y2, start=270, extent=180, fill=fill_color, outline=fill_color)
        # Draw center rectangle
        self.create_rectangle(x1 + r, y1, x2 - r, y2, fill=fill_color, outline=fill_color, width=0)

        # Draw slider
        slider_margin = 3
        if self.is_on:
            slider_x = toggle_x + self.toggle_width - self.slider_size - slider_margin
        else:
            slider_x = toggle_x + slider_margin

        slider_y = toggle_y + slider_margin

        self.create_oval(slider_x, slider_y,
                         slider_x + self.slider_size,
                         slider_y + self.slider_size,
                         fill=self.slider_color, outline='', width=0)

        # Draw text
        text_x = toggle_x + self.toggle_width + 15
        text_color = self.text_on_color if self.is_on else self.text_color
        self.create_text(text_x, 15, text=self.text, anchor='w',
                         fill=text_color, font=('Segoe UI', 10))
    
    def toggle(self, event=None):
        """Toggle the button state"""
        self.variable.set(not self.variable.get())
    
    def on_enter(self, event):
        """Handle mouse enter"""
        self.configure(cursor='hand2')
    
    def on_leave(self, event):
        """Handle mouse leave"""
        self.configure(cursor='')
    
    def on_variable_change(self, *args):
        """Handle variable change"""
        self.draw()

class VehicleGPSDecoder:
    def __init__(self, root):
        logger.info("Initializing VehicleGPSDecoder GUI")
        self.root = root
        self.root.title(f"FENDER v{FENDER_VERSION}")
        self.root.geometry("1200x700")
        self.root.configure(bg='#1a1a1a')
        self.processing_start_time = None
        self.execution_mode = "GUI"
        self.filter_duplicates = tk.BooleanVar(value=False)
        
        logger.debug("Setting up window properties")
        
        # Initialize decoder registry
        logger.info("Initializing decoder registry")
        self.decoder_registry = DecoderRegistry()
        decoder_names = self.decoder_registry.get_decoder_names()

        # Check if decoders were found
        if not decoder_names:
            logger.critical("No decoders found during initialization")
            self.root.withdraw()
            messagebox.showerror("Initialization Error",
                                 "No decoders found.\n\nPlease ensure decoder files are properly included.")
            self.root.destroy()
            return

        logger.info(f"Found {len(decoder_names)} decoders: {decoder_names}")
        
        self.current_decoder = None
        self.selected_decoder_name = decoder_names[0]
        self.decoder_buttons = {}
        self.export_format = tk.StringVar(value="xlsx")  # Default to XLSX

        self.setup_styles()
        self.input_file = None
        self.is_processing = False
        self.stop_event = threading.Event()

        self.setup_ui()
        self.setup_drag_drop()
        
        logger.info("GUI initialization complete")
    
    def decoder_supports_folders(self, decoder_name: str) -> bool:
        """Check if the decoder supports folder input instead of files"""
        logger.debug(f"Checking if decoder supports folders: {decoder_name}")
    
        try:
            decoder_class = self.decoder_registry.get_decoder(decoder_name)
            decoder_instance = decoder_class()
            # Check if get_supported_extensions returns empty list (indicates folder support)
            extensions = decoder_instance.get_supported_extensions()
            supports_folders = len(extensions) == 0
            logger.debug(f"Decoder {decoder_name} supports folders: {supports_folders}")
            return supports_folders
        except Exception as e:
            logger.error(f"Error checking folder support for {decoder_name}: {e}")
            return False

    def select_decoder(self, decoder_name: str):
        """Handle decoder selection from the button list."""
        logger.info(f"Selecting decoder: {decoder_name}")
        self.selected_decoder_name = decoder_name
        
        # Update button styles
        for name, button in self.decoder_buttons.items():
            if name == decoder_name:
                button.configure(style='Selected.TButton')
                logger.debug(f"Highlighted button for: {name}")
            else:
                button.configure(style='DecoderList.TButton')
        
        # Trigger updates
        self.on_decoder_changed()

    def setup_styles(self):
        logger.debug("Setting up GUI styles")
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Configure styles
        self.style.configure('Title.TLabel', 
                           background='#1a1a1a', 
                           foreground='#ffffff', 
                           font=('Segoe UI', 24, 'bold'))
        
        self.style.configure('Subtitle.TLabel', 
                   background='#1a1a1a', 
                   foreground='#cccccc', 
                   font=('Segoe UI', 13))
        
        self.style.configure('Dark.TFrame', 
                           background='#1a1a1a', 
                           relief='flat')
        
        self.style.configure('DropZone.TFrame', 
                           background='#252525',
                           relief='solid', 
                           borderwidth=1,
                           bordercolor='#333333')
        
        self.style.configure('Dark.TButton',
                   background='#4a9eff',
                   foreground='white',
                   font=('Segoe UI', 11),
                   borderwidth=0,
                   focuscolor='none',
                   padding=(12, 8))
        
        self.style.map('Dark.TButton',
                      background=[('active', '#3d8ce6'),
                                ('pressed', '#2d7acc')])
        
        self.style.configure('Disabled.TButton',
                   background='#888888',
                   foreground='#cccccc',
                   font=('Segoe UI', 11),
                   borderwidth=0,
                   focuscolor='none',
                   padding=(12, 8))
        
        self.style.map('Disabled.TButton',
                      background=[('active', '#888888'), ('disabled', '#888888')],
                      foreground=[('disabled', '#cccccc')])
        
        self.style.configure('Progress.TProgressbar',
                           background='#4a9eff',
                           troughcolor='#333333',
                           borderwidth=0,
                           lightcolor='#4a9eff',
                           darkcolor='#4a9eff')
        
        self.style.configure('Horizontal.TProgressbar',
                           background='#4a9eff',
                           troughcolor='#333333',
                           borderwidth=0,
                           lightcolor='#4a9eff',
                           darkcolor='#4a9eff')

        self.style.configure('DecoderList.TButton',
                   background='#252525',
                   foreground='#cccccc',
                   font=('Segoe UI', 12),
                   borderwidth=1,
                   focuscolor='none',
                   bordercolor='#1a1a1a',
                   padding=(15, 10))
        
        self.style.map('DecoderList.TButton',
                      background=[('active', '#333333')])
        
        self.style.configure('Selected.TButton',
                           background='#4a9eff',
                           foreground='white',
                           font=('Segoe UI', 11, 'bold'),
                           borderwidth=1,
                           bordercolor='#1a1a1a',
                           focuscolor='none',
                           padding=(15, 10))
        
        self.style.map('Selected.TButton',
                      background=[('active', '#3d8ce6')])
        
        logger.debug("GUI styles configured")
    
    def setup_ui(self):
        logger.info("Setting up main UI components")
        
        # Main container
        main_frame = ttk.Frame(self.root, style='Dark.TFrame')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Left Panel for Decoder Selection
        logger.debug("Creating left panel for decoder selection")
        left_panel = ttk.Frame(main_frame, style='Dark.TFrame', width=190)
        left_panel.pack_propagate(False)
        left_panel.pack(side='left', fill='y', padx=(0, 20))
        
        decoder_label = ttk.Label(left_panel, text="Select Decoder",
                                  background='#1a1a1a', foreground='#ffffff',
                                  font=('Segoe UI', 14, 'bold'))
        decoder_label.pack(anchor='w', pady=(10, 15))
        
        # Scrollable area for decoder buttons
        canvas = tk.Canvas(left_panel, bg='#1a1a1a', highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, style='Dark.TFrame')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Populate decoder buttons
        logger.debug(f"Creating buttons for {len(self.decoder_registry.get_decoder_names())} decoders")
        for decoder_name in self.decoder_registry.get_decoder_names():
            btn = ttk.Button(scrollable_frame, text=decoder_name,
                             style='DecoderList.TButton',
                             command=lambda name=decoder_name: self.select_decoder(name))
            btn.pack(fill='x', expand=True, pady=2)
            self.decoder_buttons[decoder_name] = btn
            logger.debug(f"Created button for decoder: {decoder_name}")
            
        # Right Panel for Main Content
        logger.debug("Creating right panel for main content")
        right_panel = ttk.Frame(main_frame, style='Dark.TFrame')
        right_panel.pack(side='right', fill='both', expand=True)
        
        # Header
        header_frame = ttk.Frame(right_panel, style='Dark.TFrame')
        header_frame.pack(fill='x', pady=(0, 20))
        title_label = ttk.Label(header_frame, text="Forensic Extraction of Navigational Data & Event Records", style='Title.TLabel')
        title_label.pack(anchor='w')
        subtitle_label = ttk.Label(header_frame, text="Extract and decode GPS data with timestamps from vehicle telematics binary files", style='Subtitle.TLabel')
        subtitle_label.pack(anchor='w', pady=(5, 0))

        # Export format selection with custom radio buttons
        logger.debug("Creating export format selection")
        export_frame = ttk.Frame(right_panel, style='Dark.TFrame')
        export_frame.pack(fill='x', pady=(0, 15))
        
        export_label = ttk.Label(export_frame, text="Export Format:",
                                background='#1a1a1a', foreground='#ffffff',
                                font=('Segoe UI', 12, 'bold'))
        export_label.pack(anchor='w', pady=(0, 5))
        
        format_frame = ttk.Frame(export_frame, style='Dark.TFrame')
        format_frame.pack(anchor='w')
        
        # Custom radio buttons for export format
        xlsx_radio = CustomRadiobutton(format_frame, "Excel (.xlsx)", 
                                      self.export_format, "xlsx",
                                      bg='#1a1a1a')
        xlsx_radio.pack(side='left', padx=(0, 20))
        
        csv_radio = CustomRadiobutton(format_frame, "CSV (.csv)", 
                                     self.export_format, "csv",
                                     bg='#1a1a1a')
        csv_radio.pack(side='left', padx=(0, 20))
        
        json_radio = CustomRadiobutton(format_frame, "JSON (.json)", 
                                      self.export_format, "json",
                                      bg='#1a1a1a')
        json_radio.pack(side='left', padx=(0, 20))
        
        geojson_radio = CustomRadiobutton(format_frame, "GeoJSON (.geojson)", 
                                         self.export_format, "geojson",
                                         bg='#1a1a1a')
        geojson_radio.pack(side='left')

        kml_radio = CustomRadiobutton(format_frame, "KML (.kml)", 
                             self.export_format, "kml",
                             bg='#1a1a1a')
        kml_radio.pack(side='left')

        # Filter controls
        filter_frame = ttk.Frame(right_panel, style='Dark.TFrame')
        filter_frame.pack(fill='x', pady=(15, 15))

        filter_label = ttk.Label(filter_frame, text="Filtering Options:",
                                background='#1a1a1a', foreground='#ffffff',
                                font=('Segoe UI', 12, 'bold'))
        filter_label.pack(anchor='w', pady=(0, 5))

        toggle_frame = ttk.Frame(filter_frame, style='Dark.TFrame')
        toggle_frame.pack(anchor='w')

        # Filter duplicates toggle
        filter_toggle = CustomToggleButton(toggle_frame, "Filter Duplicate Entries",
                                         self.filter_duplicates,
                                         bg='#1a1a1a')
        filter_toggle.pack(side='left', padx=(0, 20))

        # Info label
        info_label = ttk.Label(filter_frame, 
                      text=f"When enabled, filter GPS entries with identical timestamps and coordinates within {decimals_of_prec} decimal places",
                      background='#1a1a1a', foreground='#888888',
                      font=('Segoe UI', 9))
        info_label.pack(anchor='w', pady=(5, 0))

        # Drop zone
        logger.debug("Creating file drop zone")
        self.drop_frame = ttk.Frame(right_panel, style='DropZone.TFrame')
        self.drop_frame.pack(fill='both', expand=True, pady=(0, 20))
        drop_content = tk.Frame(self.drop_frame, bg='#252525')
        drop_content.place(relx=0.5, rely=0.5, anchor='center')
        icon_label = tk.Label(drop_content, text="📁", bg='#252525', fg='#4a9eff', font=('Segoe UI', 48))
        icon_label.pack(pady=(0, 10))
        self.drop_label = tk.Label(drop_content, text="", bg='#252525', fg='#cccccc', font=('Segoe UI', 14), justify='center')
        self.drop_label.pack()
        self.file_info_label = tk.Label(drop_content, text="", bg='#252525', fg='#888888', font=('Segoe UI', 10))
        self.file_info_label.pack(pady=(10, 0))

        # Buttons frame
        logger.debug("Creating control buttons")
        button_frame = ttk.Frame(right_panel, style='Dark.TFrame')
        button_frame.pack(fill='x', pady=(0, 20))
        self.browse_btn = ttk.Button(button_frame, text="Browse Files", style='Dark.TButton', command=self.browse_file)
        self.browse_btn.pack(side='left', padx=(0, 10))
        self.process_btn = ttk.Button(button_frame, text="Process File", style='Disabled.TButton', command=self.process_file, state='disabled')
        self.process_btn.pack(side='left')
        self.clear_btn = ttk.Button(button_frame, text="Clear", style='Disabled.TButton', command=self.clear_file, state='disabled')
        self.clear_btn.pack(side='right')
        self.stop_btn = ttk.Button(button_frame, text="Stop Processing", style='Disabled.TButton', command=self.stop_processing, state='disabled')
        self.stop_btn.pack(side='right', padx=(0, 10))

        # Progress section
        logger.debug("Creating progress indicators")
        progress_frame = ttk.Frame(right_panel, style='Dark.TFrame')
        progress_frame.pack(fill='x')
        self.progress_label = ttk.Label(progress_frame, text="", background='#1a1a1a', foreground='#cccccc', font=('Segoe UI', 10))
        self.progress_label.pack(anchor='w', pady=(0, 5))
        self.progress = ttk.Progressbar(progress_frame, style='Horizontal.TProgressbar', mode='determinate', length=300)
        self.progress.pack(fill='x')

        # Results section
        logger.debug("Creating results display")
        results_frame = ttk.Frame(right_panel, style='Dark.TFrame')
        results_frame.pack(fill='x', pady=(20, 0))
        self.results_label = ttk.Label(results_frame, text="", background='#1a1a1a', foreground='#4a9eff', font=('Segoe UI', 11, 'bold'))
        self.results_label.pack(anchor='w')

        # Set initial state
        self.select_decoder(self.selected_decoder_name)
        
        logger.info("UI setup complete")
    
    def setup_drag_drop(self):
        logger.debug("Setting up drag and drop functionality")
        
        # Bind click event to drop zone
        self.drop_frame.bind("<Button-1>", lambda e: self.browse_file())
        self.drop_label.bind("<Button-1>", lambda e: self.browse_file())
        
        # Enable drag-and-drop
        self.drop_frame.drop_target_register(DND_FILES)
        self.drop_frame.dnd_bind('<<Drop>>', self.on_file_drop)
        
        logger.debug("Drag and drop setup complete")
    
    def add_filter_controls(self):
        """Add filter controls to the UI - call this in setup_ui()"""
        # This should be added after the export format selection in setup_ui()
    
        # Filter frame
        filter_frame = ttk.Frame(self.right_panel, style='Dark.TFrame')
        filter_frame.pack(fill='x', pady=(15, 15))
    
        # Filter label
        filter_label = ttk.Label(filter_frame, text="Filtering Options:",
                                background='#1a1a1a', foreground='#ffffff',
                                font=('Segoe UI', 12, 'bold'))
        filter_label.pack(anchor='w', pady=(0, 5))
    
        # Toggle and precision frame
        toggle_frame = ttk.Frame(filter_frame, style='Dark.TFrame')
        toggle_frame.pack(anchor='w')
    
        # Filter duplicates toggle
        self.filter_duplicates = tk.BooleanVar(value=False)
        filter_toggle = CustomToggleButton(toggle_frame, "Filter Duplicate Entries",
                                         self.filter_duplicates,
                                         bg='#1a1a1a')
        filter_toggle.pack(side='left', padx=(0, 20))
    
        # Precision selection
        precision_label = ttk.Label(toggle_frame, text="Precision:",
                                   background='#1a1a1a', foreground='#cccccc',
                                   font=('Segoe UI', 10))
        precision_label.pack(side='left', padx=(0, 5))
    
        # Precision spinbox
        self.precision_var = tk.IntVar(value=4)
        precision_spin = ttk.Spinbox(toggle_frame, from_=1, to=8,
                                    textvariable=self.precision_var,
                                    width=5, font=('Segoe UI', 10))
        precision_spin.pack(side='left', padx=(0, 5))
    
        decimals_label = ttk.Label(toggle_frame, text="decimal places",
                                  background='#1a1a1a', foreground='#cccccc',
                                  font=('Segoe UI', 10))
        decimals_label.pack(side='left')
    
        # Info label
        info_label = ttk.Label(filter_frame, 
                              text="When enabled, removes GPS entries with identical timestamps and coordinates within the specified precision",
                              background='#1a1a1a', foreground='#888888',
                              font=('Segoe UI', 9))
        info_label.pack(anchor='w', pady=(5, 0))


    def apply_duplicate_filter(self, entries: List['GPSEntry']) -> List['GPSEntry']:
        """Apply duplicate filtering if enabled"""
        if self.filter_duplicates.get():
            precision = decimals_of_prec
            logger.info(f"Applying duplicate filter with precision={precision}")
            return filter_duplicate_entries(entries, precision, logger)
        else:
            logger.info("Duplicate filtering is disabled")
            return entries


    # Modified process_in_background method (update the part after entries are extracted):
    def process_in_background_with_filter(self, input_path, output_path):
        """Modified version of process_in_background that includes filtering"""
        logger.info(f"Background processing started for: {input_path}")
    
        def progress_callback(status, percent):
            logger.debug(f"Progress update: {status} ({percent}%)")
            self.root.after(0, self.update_progress, status, percent)

        try:
            # Pass stop_event to decoder
            logger.info("Calling decoder extract_gps_data method")
            entries, error = self.current_decoder.extract_gps_data(
                input_path, progress_callback, stop_event=self.stop_event
            )
        
            if error:
                logger.error(f"Decoder returned error: {error}")
                self.root.after(0, self.processing_error, error)
            else:
                logger.info(f"Decoder extracted {len(entries)} entries")
            
                # Apply duplicate filter if enabled
                filtered_entries = self.apply_duplicate_filter(entries)
            
                # Update progress if filtering was applied
                if len(filtered_entries) < len(entries):
                    self.root.after(0, self.update_progress, 
                                   f"Filtered {len(entries) - len(filtered_entries)} duplicates...", 80)
            
                # Write to selected format
                format_type = self.export_format.get()
                logger.info(f"Writing output in {format_type} format")
                self.root.after(0, self.update_progress, f"Writing {format_type.upper()} file...", 85)
            
                if format_type == "xlsx":
                    self.write_xlsx(filtered_entries, output_path)
                elif format_type == "csv":
                    self.write_csv(filtered_entries, output_path)
                elif format_type == "json":
                    self.write_json(filtered_entries, output_path)
                elif format_type == "geojson":
                    write_geojson(filtered_entries, output_path, self.selected_decoder_name)
                # In the write output section, add this after the geojson condition:
                elif format_type == "kml":
                    write_kml(filtered_entries, output_path, self.selected_decoder_name)
            
                # Report both original and filtered counts if filtering was applied
                if self.filter_duplicates.get() and len(filtered_entries) < len(entries):
                    result_info = {
                        'original_count': len(entries),
                        'filtered_count': len(filtered_entries),
                        'duplicates_removed': len(entries) - len(filtered_entries)
                    }
                    self.root.after(0, self.processing_complete_with_filter_info, 
                                  result_info, output_path)
                else:
                    self.root.after(0, self.processing_complete, len(filtered_entries), output_path)
                
        except Exception as e:
            logger.error(f"Exception during background processing: {e}", exc_info=True)
            self.root.after(0, self.processing_error, str(e))


    def processing_complete_with_filter_info(self, result_info, output_path):
        """Modified processing complete handler that shows filtering information"""
        logger.info(f"Processing completed with filtering. Original: {result_info['original_count']}, "
                    f"Filtered: {result_info['filtered_count']}, Output: {output_path}")
    
        self.is_processing = False
        self.process_btn.configure(state='normal', text='Process File', style='Dark.TButton')
        self.browse_btn.configure(state='normal')
        self.clear_btn.configure(state='normal', style='Dark.TButton')
        self.stop_btn.configure(state='disabled', style='Disabled.TButton')
    
        self.progress_label.configure(text="Processing complete!")
        self.progress['value'] = 100
    
        output_filename = os.path.basename(output_path)
        format_type = self.export_format.get().upper()
    
        result_text = (f"Successfully extracted {result_info['filtered_count']} GPS entries to {format_type}:\n"
                       f" {output_filename}\n"
                       f" ({result_info['duplicates_removed']} duplicates removed from {result_info['original_count']} total entries)")
    
        self.results_label.configure(text=result_text)
    
        processing_time = (datetime.now() - self.processing_start_time).total_seconds() if self.processing_start_time else 0
        logger.info(f"Total processing time: {processing_time:.2f} seconds")

    def on_decoder_changed(self):
        """Handle decoder type change"""
        logger.info(f"Decoder changed to: {self.selected_decoder_name}")
        
        decoder_class = self.decoder_registry.get_decoder(self.selected_decoder_name)
        decoder_instance = decoder_class()
        dropzone_text = decoder_instance.get_dropzone_text()
        self.drop_label.configure(text=dropzone_text)
        logger.debug(f"Updated dropzone text: {dropzone_text}")

        if self.input_file:
            logger.debug("Clearing current file due to decoder change")
            self.clear_file()

        # Update browse button text based on decoder type
        if self.decoder_supports_folders(self.selected_decoder_name):
            self.browse_btn.configure(text="Browse Folders")
        else:
            self.browse_btn.configure(text="Browse Files")
    
    def browse_file(self):
        logger.info("Browse dialog opened")
    
        if self.is_processing:
            logger.warning("Browse attempted while processing")
            return
    
        # Check if decoder supports folders
        if self.decoder_supports_folders(self.selected_decoder_name):
            logger.info(f"Decoder {self.selected_decoder_name} requires folder selection")
        
            folder_path = filedialog.askdirectory(
                title=f"Select {self.selected_decoder_name} Data Folder"
            )
        
            if folder_path:
                logger.info(f"Folder selected: {folder_path}")
                self.set_input_file(folder_path)
            else:
                logger.debug("Folder selection cancelled")
        else:
            # Original file selection logic
            decoder_class = self.decoder_registry.get_decoder(self.selected_decoder_name)
            decoder_instance = decoder_class()
            extensions = decoder_instance.get_supported_extensions()
            logger.debug(f"Supported extensions for {self.selected_decoder_name}: {extensions}")
        
            filetypes = []
            if extensions:
                ext_str = ";".join([f"*{ext}" for ext in extensions])
                filetypes.append((f"{self.selected_decoder_name} files", ext_str))
            filetypes.append(("All files", "*.*"))
        
            file_path = filedialog.askopenfilename(
                title=f"Select {self.selected_decoder_name} Binary File",
                filetypes=filetypes
            )
        
            if file_path:
                logger.info(f"File selected: {file_path}")
                # Validate file path
                is_valid, result = validate_file_path(file_path, extensions)
                if is_valid:
                    self.set_input_file(result)
                else:
                    logger.error(f"File validation failed: {result}")
                    messagebox.showerror("File Validation Error", result)
            else:
                logger.debug("File selection cancelled")
    
    def set_input_file(self, file_path):
        logger.info(f"Setting input path: {file_path}")
    
        self.input_file = file_path
    
        # Check if it's a folder or file
        if os.path.isdir(file_path):
            folder_name = os.path.basename(file_path)
            # Count files in folder for size estimation
            total_size = 0
            file_count = 0
            for root, dirs, files in os.walk(file_path):
                for file in files:
                    file_count += 1
                    try:
                        total_size += os.path.getsize(os.path.join(root, file))
                    except:
                        pass
        
            size_mb = total_size / (1024 * 1024)
            logger.debug(f"Folder details - Name: {folder_name}, Files: {file_count}, Total size: {size_mb:.2f} MB")
        
            self.drop_label.configure(text=f"Selected Folder: {folder_name}")
            self.file_info_label.configure(text=f"Contains {file_count} files, Total size: {size_mb:.2f} MB")
        else:
            filename = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            size_mb = file_size / (1024 * 1024)
        
            logger.debug(f"File details - Name: {filename}, Size: {size_mb:.2f} MB")
        
            self.drop_label.configure(text=f"Selected: {filename}")
            self.file_info_label.configure(text=f"Size: {size_mb:.2f} MB")
    
        self.process_btn.configure(state='normal', style='Dark.TButton')
        self.clear_btn.configure(state='normal', style='Dark.TButton')
    
        logger.info("Input path set successfully")
    
    def clear_file(self):
        logger.info("Clearing current file")
        
        if self.is_processing:
            logger.warning("Clear attempted while processing")
            return
        
        self.input_file = None
        decoder_class = self.decoder_registry.get_decoder(self.selected_decoder_name)
        decoder_instance = decoder_class()
        self.drop_label.configure(text=decoder_instance.get_dropzone_text())
        self.file_info_label.configure(text="")
        self.process_btn.configure(state='disabled', style='Disabled.TButton')
        self.progress_label.configure(text="")
        self.progress['value'] = 0
        self.results_label.configure(text="")
        self.clear_btn.configure(state='disabled', style='Disabled.TButton')
        
        logger.info("File cleared")
    
    def generate_timestamped_filename(self, base_path: str, decoder_name: str, format_ext: str) -> str:
        """Generate filename with timestamp"""
        base, _ = os.path.splitext(base_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_decoder_name = sanitize_filename(decoder_name)
        filename = f"{base}_{safe_decoder_name}_{timestamp}.{format_ext}"
        
        logger.debug(f"Generated timestamped filename: {filename}")
        return filename
    
    def process_file(self):
        logger.info("Starting file processing")
        
        if not self.input_file or self.is_processing:
            logger.warning("Process attempted with no file or already processing")
            return
    
        self.stop_event.clear()
        
        self.processing_start_time = datetime.now()
        logger.debug(f"Processing start time: {self.processing_start_time}")
        
        self.is_processing = True
        self.process_btn.configure(state='disabled', text='Processing...')
        self.browse_btn.configure(state='disabled')
        self.clear_btn.configure(state='disabled')
        self.stop_btn.configure(state='normal', style='Dark.TButton')
    
        # Get decoder
        decoder_class = self.decoder_registry.get_decoder(self.selected_decoder_name)
        self.current_decoder = decoder_class()
        logger.info(f"Using decoder: {self.selected_decoder_name}")
    
        # Generate output path with timestamp and selected format
        format_ext = self.export_format.get()
        output_path = self.generate_timestamped_filename(self.input_file, self.selected_decoder_name, format_ext)
        logger.info(f"Output will be saved to: {output_path}")
    
        # Start processing in a separate thread
        thread = threading.Thread(target=self.process_in_background_with_filter,
            args=(self.input_file, output_path))
        thread.daemon = True
        thread.start()
        logger.debug("Started processing thread")
    
    def process_in_background(self, input_path, output_path):
        logger.info(f"Background processing started for: {input_path}")
        
        def progress_callback(status, percent):
            logger.debug(f"Progress update: {status} ({percent}%)")
            self.root.after(0, self.update_progress, status, percent)

        try:
            # Pass stop_event to decoder
            logger.info("Calling decoder extract_gps_data method")
            entries, error = self.current_decoder.extract_gps_data(
                input_path, progress_callback, stop_event=self.stop_event
            )
            
            if error:
                logger.error(f"Decoder returned error: {error}")
                self.root.after(0, self.processing_error, error)
            else:
                logger.info(f"Decoder extracted {len(entries)} entries")
                
                # Write to selected format
                format_type = self.export_format.get()
                logger.info(f"Writing output in {format_type} format")
                self.root.after(0, self.update_progress, f"Writing {format_type.upper()} file...", 85)
                
                if format_type == "xlsx":
                    self.write_xlsx(entries, output_path)
                elif format_type == "csv":
                    self.write_csv(entries, output_path)
                elif format_type == "json":
                    self.write_json(entries, output_path)
                elif format_type == "geojson":
                    write_geojson(entries, output_path, self.selected_decoder_name)
                # Add this after the geojson condition:
                elif format_type == "kml":
                    write_kml(entries, output_path, self.selected_decoder_name)
                
                self.root.after(0, self.processing_complete, len(entries), output_path)
                
        except Exception as e:
            logger.error(f"Exception during background processing: {e}", exc_info=True)
            self.root.after(0, self.processing_error, str(e))
    
    def write_xlsx(self, entries: List[GPSEntry], output_path: str):
        """Write GPS entries to XLSX file using decoder-specific format with extraction details"""
        logger.info(f"Writing {len(entries)} entries to XLSX file: {output_path}")
        
        wb = Workbook()
    
        # Main GPS Data worksheet
        ws_data = wb.active
        ws_data.title = "GPS Data"
    
        # Get headers from decoder
        headers = self.current_decoder.get_xlsx_headers()
        ws_data.append(headers)
        logger.debug(f"XLSX headers: {headers}")
    
        # Write entries
        for i, entry in enumerate(entries):
            row = self.current_decoder.format_entry_for_xlsx(entry)
            ws_data.append(row)
            if i % 100 == 0:
                logger.debug(f"Written {i} entries to XLSX")
    
        # Create Extraction Details worksheet
        logger.debug("Creating Extraction Details worksheet")
        ws_details = wb.create_sheet("Extraction Details")
    
        # Get system and extraction info
        system_info = get_system_info(
            input_file=self.input_file,
            output_file=output_path,
            execution_mode=self.execution_mode,
            decoder_registry=self.decoder_registry
        )
        processing_time = (datetime.now() - self.processing_start_time).total_seconds() if self.processing_start_time else None
        extraction_info = get_extraction_info(
            self.selected_decoder_name, 
            self.input_file, 
            output_path, 
            len(entries),
            processing_time
        )
    
        # Write extraction details
        ws_details.append(["FENDER Extraction Report"])
        ws_details.append([])
    
        # System Information
        ws_details.append(["System Information"])
        ws_details.append(["Field", "Value"])
        for key, value in system_info.items():
            if key != "decoder_hashes":  # Handle separately
                ws_details.append([key.replace("_", " ").title(), str(value)])
    
        ws_details.append([])
    
        # Decoder Hashes
        if "decoder_hashes" in system_info:
            ws_details.append(["Decoder Integrity Verification"])
            ws_details.append(["Decoder", "File Path", "SHA256 Hash", "File Size", "Last Modified"])
            for decoder_name, hash_info in system_info["decoder_hashes"].items():
                if "error" in hash_info:
                    ws_details.append([decoder_name, "Error", hash_info["error"], "", ""])
                else:
                    ws_details.append([
                        decoder_name,
                        hash_info["file_path"],
                        hash_info["sha256_hash"],
                        hash_info["file_size"],
                        hash_info["last_modified"]
                    ])
    
        ws_details.append([])
    
        # Extraction Information
        ws_details.append(["Extraction Information"])
        ws_details.append(["Field", "Value"])
    
        # Input file details
        ws_details.append(["Input File Path", extraction_info["input_file"]["path"]])
        ws_details.append(["Input File Name", extraction_info["input_file"]["filename"]])
        ws_details.append(["Input File Size (MB)", extraction_info["input_file"]["size_mb"]])
        ws_details.append(["Input File SHA256", extraction_info["input_file"]["sha256_hash"]])
    
        # Output file details
        ws_details.append(["Output File Path", extraction_info["output_file"]["path"]])
        ws_details.append(["Output File Name", extraction_info["output_file"]["filename"]])
    
        # Extraction details
        ws_details.append(["Decoder Used", extraction_info["extraction_details"]["decoder_used"]])
        ws_details.append(["Entries Extracted", extraction_info["extraction_details"]["entries_extracted"]])
        if processing_time:
            ws_details.append(["Processing Time (seconds)", round(processing_time, 2)])
    
        # Format the details worksheet
        ws_details.column_dimensions['A'].width = 25
        ws_details.column_dimensions['B'].width = 50
        ws_details.column_dimensions['C'].width = 70
    
        logger.debug("Saving XLSX file")
        wb.save(output_path)
        logger.info(f"XLSX file saved successfully: {output_path}")
    
    def write_csv(self, entries: List[GPSEntry], output_path: str):
        """Write GPS entries to CSV file with extraction details"""
        logger.info(f"Writing {len(entries)} entries to CSV file: {output_path}")
        
        headers = self.current_decoder.get_xlsx_headers()
    
        # Get system and extraction info
        system_info = get_system_info(
            input_file=self.input_file,
            output_file=output_path,
            execution_mode=self.execution_mode,
            decoder_registry=self.decoder_registry
        )
        processing_time = (datetime.now() - self.processing_start_time).total_seconds() if self.processing_start_time else None
        extraction_info = get_extraction_info(
            self.selected_decoder_name, 
            self.input_file, 
            output_path, 
            len(entries),
            processing_time
        )
    
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
        
            # Write headers
            writer.writerow(headers)
        
            # Write entries
            for i, entry in enumerate(entries):
                row = self.current_decoder.format_entry_for_xlsx(entry)
                writer.writerow(row)
                if i % 100 == 0:
                    logger.debug(f"Written {i} entries to CSV")
        
            # Add separator
            for _ in range(5):
                writer.writerow([])
        
            # Write extraction details
            writer.writerow(["FENDER Extraction Report"])
            writer.writerow([])
        
            # System Information
            writer.writerow(["System Information"])
            writer.writerow(["Field", "Value"])
            for key, value in system_info.items():
                if key != "decoder_hashes":
                    writer.writerow([key.replace("_", " ").title(), str(value)])
        
            writer.writerow([])
        
            # Decoder Hashes
            if "decoder_hashes" in system_info:
                writer.writerow(["Decoder Integrity Verification"])
                writer.writerow(["Decoder", "File Path", "SHA256 Hash", "File Size", "Last Modified"])
                for decoder_name, hash_info in system_info["decoder_hashes"].items():
                    if "error" in hash_info:
                        writer.writerow([decoder_name, "Error", hash_info["error"], "", ""])
                    else:
                        writer.writerow([
                            decoder_name,
                            hash_info["file_path"],
                            hash_info["sha256_hash"],
                            hash_info["file_size"],
                            hash_info["last_modified"]
                        ])
        
            writer.writerow([])
        
            # Extraction Information
            writer.writerow(["Extraction Information"])
            writer.writerow(["Field", "Value"])
        
            # Input file details
            writer.writerow(["Input File Path", extraction_info["input_file"]["path"]])
            writer.writerow(["Input File Name", extraction_info["input_file"]["filename"]])
            writer.writerow(["Input File Size (MB)", extraction_info["input_file"]["size_mb"]])
            writer.writerow(["Input File SHA256", extraction_info["input_file"]["sha256_hash"]])
        
            # Output file details
            writer.writerow(["Output File Path", extraction_info["output_file"]["path"]])
            writer.writerow(["Output File Name", extraction_info["output_file"]["filename"]])
        
            # Extraction details
            writer.writerow(["Decoder Used", extraction_info["extraction_details"]["decoder_used"]])
            writer.writerow(["Entries Extracted", extraction_info["extraction_details"]["entries_extracted"]])
            if processing_time:
                writer.writerow(["Processing Time (seconds)", round(processing_time, 2)])
                
        logger.info(f"CSV file saved successfully: {output_path}")
    
    def write_json(self, entries: List[GPSEntry], output_path: str):
        """Write GPS entries to JSON file with extraction details"""
        logger.info(f"Writing {len(entries)} entries to JSON file: {output_path}")
        
        # Get system and extraction info
        system_info = get_system_info(
            input_file=self.input_file,
            output_file=output_path,
            execution_mode=self.execution_mode,
            decoder_registry=self.decoder_registry
        )
        processing_time = (datetime.now() - self.processing_start_time).total_seconds() if self.processing_start_time else None
        extraction_info = get_extraction_info(
            self.selected_decoder_name, 
            self.input_file, 
            output_path, 
            len(entries),
            processing_time
        )
    
        json_data = {
            "metadata": {
                "decoder": self.selected_decoder_name,
                "extraction_timestamp": datetime.now().isoformat(),
                "total_entries": len(entries)
            },
            "system_information": system_info,
            "extraction_information": extraction_info,
            "gps_entries": []
        }
    
        headers = self.current_decoder.get_xlsx_headers()
    
        for i, entry in enumerate(entries):
            row = self.current_decoder.format_entry_for_xlsx(entry)
            entry_dict = {}
        
            # Map row data to headers
            for j, header in enumerate(headers):
                if j < len(row):
                    entry_dict[header] = row[j]
        
            # Add core GPS data
            entry_dict.update({
                "latitude": entry.latitude,
                "longitude": entry.longitude,
                "timestamp": entry.timestamp,
                "extra_data": entry.extra_data
            })
        
            json_data["gps_entries"].append(entry_dict)
            
            if i % 100 == 0:
                logger.debug(f"Processed {i} entries for JSON")
    
        logger.debug("Writing JSON to file")
        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False, default=str)
            
        logger.info(f"JSON file saved successfully: {output_path}")
    
    def update_progress(self, status, percent):
        logger.debug(f"UI progress update: {status} ({percent}%)")
        self.progress_label.configure(text=status)
        self.progress['value'] = percent
        self.root.update_idletasks()
    
    def processing_complete(self, entry_count, output_path):
        logger.info(f"Processing completed successfully. Entries: {entry_count}, Output: {output_path}")
        
        self.is_processing = False
        self.process_btn.configure(state='normal', text='Process File', style='Dark.TButton')
        self.browse_btn.configure(state='normal')
        self.clear_btn.configure(state='normal', style='Dark.TButton')
        self.stop_btn.configure(state='disabled', style='Disabled.TButton')
        
        self.progress_label.configure(text="Processing complete!")
        self.progress['value'] = 100
        
        output_filename = os.path.basename(output_path)
        format_type = self.export_format.get().upper()
        result_text = f"Successfully extracted {entry_count} GPS entries to {format_type}:\n {output_filename}"
        
        self.results_label.configure(text=result_text)
        
        processing_time = (datetime.now() - self.processing_start_time).total_seconds() if self.processing_start_time else 0
        logger.info(f"Total processing time: {processing_time:.2f} seconds")
    
    def processing_error(self, error_msg):
        logger.error(f"Processing failed with error: {error_msg}")
        
        self.is_processing = False
        self.process_btn.configure(state='normal', text='Process File')
        self.browse_btn.configure(state='normal')
        self.clear_btn.configure(state='normal')
        self.stop_btn.configure(state='disabled', style='Disabled.TButton')
        
        self.progress_label.configure(text="Processing failed!")
        self.progress['value'] = 0
        self.results_label.configure(text=f"Error: {error_msg}")
        
        messagebox.showerror("Processing Error", f"Failed to process file:\n\n{error_msg}")
    
    def stop_processing(self):
        logger.info("Stop processing requested by user")
        
        if self.is_processing:
            self.stop_event.set()
            self.progress_label.configure(text="Stopping...")
            self.stop_btn.configure(state='disabled', style='Disabled.TButton')
            logger.debug("Stop event set")
    
    def on_file_drop(self, event):
        logger.info("Item dropped onto drop zone")
    
        if self.is_processing:
            logger.warning("Drop ignored - currently processing")
            return
        
        dropped_path = event.data.strip().strip('{}')  # Handle paths with spaces
        logger.debug(f"Dropped path: {dropped_path}")
    
        # Check if decoder supports folders
        if self.decoder_supports_folders(self.selected_decoder_name):
            if os.path.isdir(dropped_path):
                # Validate dropped folder
                is_valid, result = validate_folder_path(dropped_path)
                if is_valid:
                    self.set_input_file(result)
                else:
                    logger.error(f"Dropped folder validation failed: {result}")
                    messagebox.showerror("Folder Validation Error", result)
            else:
                logger.warning(f"Decoder {self.selected_decoder_name} requires a folder, not a file")
                messagebox.showwarning("Invalid Input", 
                                     f"{self.selected_decoder_name} decoder requires a folder, not a file.")
        else:
            if os.path.isfile(dropped_path):
                # Original file validation logic
                decoder_class = self.decoder_registry.get_decoder(self.selected_decoder_name)
                decoder_instance = decoder_class()
                extensions = decoder_instance.get_supported_extensions()
            
                is_valid, result = validate_file_path(dropped_path, extensions)
                if is_valid:
                    self.set_input_file(result)
                else:
                    logger.error(f"Dropped file validation failed: {result}")
                    messagebox.showerror("File Validation Error", result)
            else:
                logger.warning(f"Dropped item is not a file: {dropped_path}")
                messagebox.showwarning("Invalid Input", 
                                     f"{self.selected_decoder_name} decoder requires a file, not a folder.")

def run_cli():
    """Run the CLI version with enhanced export options"""
    logger.info("Starting FENDER in CLI mode")
    
    print("Vehicle GPS Decoder - CLI Mode")
    print("=" * 40)
    
    # Initialize registry
    registry = DecoderRegistry()
    decoder_names = registry.get_decoder_names()
    
    if not decoder_names:
        logger.error("No decoders found in CLI mode")
        print("Error: No decoders found!")
        return
    
    logger.info(f"Available decoders in CLI: {decoder_names}")
    
    # Select decoder
    print("\nAvailable decoders:")
    for i, name in enumerate(decoder_names, 1):
        print(f"{i}. {name}")
    
    while True:
        try:
            choice = int(input("\nSelect decoder (enter number): "))
            if 1 <= choice <= len(decoder_names):
                selected_decoder = decoder_names[choice - 1]
                logger.info(f"CLI decoder selected: {selected_decoder}")
                break
            else:
                print("Invalid choice. Please try again.")
        except ValueError:
            print("Please enter a valid number.")
    
    # Select export format
    # Update the export format selection:
    print("\nExport formats:")
    print("1. Excel (.xlsx)")
    print("2. CSV (.csv)")
    print("3. JSON (.json)")
    print("4. GeoJSON (.geojson)")
    print("5. KML (.kml)")

    format_map = {1: "xlsx", 2: "csv", 3: "json", 4: "geojson", 5: "kml"}
    while True:
        try:
            format_choice = int(input("\nSelect export format (enter number): "))
            if 1 <= format_choice <= 5:
                export_format = format_map[format_choice]
                logger.info(f"CLI export format selected: {export_format}")
                break
            else:
                print("Invalid choice. Please try again.")
        except ValueError:
            print("Please enter a valid number.")
    
    # Check if decoder supports folders
    decoder_class = registry.get_decoder(selected_decoder)
    decoder_instance = decoder_class()
    extensions = decoder_instance.get_supported_extensions()

    if len(extensions) == 0:  # Folder-based decoder
        input_path = input(f"\nEnter the path to the {selected_decoder} data FOLDER: ").strip()
        logger.info(f"CLI input folder: {input_path}")
    
        # Validate folder
        is_valid, result = validate_folder_path(input_path)
        if not is_valid:
            logger.error(f"CLI folder validation failed: {result}")
            print(f"Error: {result}")
            return
    
        input_file = result
    else:
        # Original file input logic
        input_file = input(f"\nEnter the path to the {selected_decoder} file: ").strip()
        logger.info(f"CLI input file: {input_file}")
    
        # Validate file
        is_valid, result = validate_file_path(input_file, extensions)
        if not is_valid:
            logger.error(f"CLI file validation failed: {result}")
            print(f"Error: {result}")
            return
    
        input_file = result
    
    input_file = result
    
    # Create decoder and process
    decoder = decoder_instance
    
    # Generate timestamped output filename
    base, _ = os.path.splitext(input_file)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_decoder_name = sanitize_filename(selected_decoder)
    output_file = f"{base}_{safe_decoder_name}_{timestamp}.{export_format}"
    
    print(f"\nProcessing {selected_decoder} file...")
    logger.info(f"Starting CLI extraction process")
    
    processing_start_time = datetime.now()
    
    def progress_callback(status, percent):
        print(f"{status} ({percent}%)")
        logger.debug(f"CLI progress: {status} ({percent}%)")
    
    entries, error = decoder.extract_gps_data(input_file, progress_callback)

    processing_time = (datetime.now() - processing_start_time).total_seconds()

    # Ask about duplicate filtering
    filter_choice = input("\nFilter duplicate entries? (y/n): ").strip().lower()
    if filter_choice == 'y':
        entries = filter_duplicate_entries(entries, decimals_of_prec, logger)
        print(f"Filtered to {len(entries)} unique entries")

    # Get system and extraction info for CLI
    system_info = get_system_info(
        input_file=input_file,
        output_file=output_file,
        execution_mode="CLI",
        decoder_registry=registry
    )
    extraction_info = get_extraction_info(selected_decoder, input_file, output_file, len(entries), processing_time)

    if error:
        logger.error(f"CLI extraction error: {error}")
        print(f"Error: {error}")
    else:
        logger.info(f"CLI extraction successful: {len(entries)} entries")
        
        # Write to selected format
        if export_format == "xlsx":
            logger.debug("Writing XLSX output")
            wb = Workbook()
            
            # Main GPS Data worksheet
            ws_data = wb.active
            ws_data.title = "GPS Data"
            
            headers = decoder.get_xlsx_headers()
            ws_data.append(headers)
            
            for entry in entries:
                row = decoder.format_entry_for_xlsx(entry)
                ws_data.append(row)
            
            # Create Extraction Details worksheet
            ws_details = wb.create_sheet("Extraction Details")
            
            # Write extraction details
            ws_details.append(["FENDER Extraction Report"])
            ws_details.append([])
            
            # System Information
            ws_details.append(["System Information"])
            ws_details.append(["Field", "Value"])
            for key, value in system_info.items():
                if key != "decoder_hashes":
                    ws_details.append([key.replace("_", " ").title(), str(value)])
            
            ws_details.append([])
            
            # Decoder Hashes
            if "decoder_hashes" in system_info:
                ws_details.append(["Decoder Integrity Verification"])
                ws_details.append(["Decoder", "File Path", "SHA256 Hash", "File Size", "Last Modified"])
                for decoder_name, hash_info in system_info["decoder_hashes"].items():
                    if "error" in hash_info:
                        ws_details.append([decoder_name, "Error", hash_info["error"], "", ""])
                    else:
                        ws_details.append([
                            decoder_name,
                            hash_info["file_path"],
                            hash_info["sha256_hash"],
                            hash_info["file_size"],
                            hash_info["last_modified"]
                        ])
            
            ws_details.append([])
            
            # Extraction Information
            ws_details.append(["Extraction Information"])
            ws_details.append(["Field", "Value"])
            
            # Input file details
            ws_details.append(["Input File Path", extraction_info["input_file"]["path"]])
            ws_details.append(["Input File Name", extraction_info["input_file"]["filename"]])
            ws_details.append(["Input File Size (MB)", extraction_info["input_file"]["size_mb"]])
            ws_details.append(["Input File SHA256", extraction_info["input_file"]["sha256_hash"]])
            
            # Output file details
            ws_details.append(["Output File Path", extraction_info["output_file"]["path"]])
            ws_details.append(["Output File Name", extraction_info["output_file"]["filename"]])
            
            # Extraction details
            ws_details.append(["Decoder Used", extraction_info["extraction_details"]["decoder_used"]])
            ws_details.append(["Entries Extracted", extraction_info["extraction_details"]["entries_extracted"]])
            ws_details.append(["Processing Time (seconds)", round(processing_time, 2)])
            
            # Format the details worksheet
            ws_details.column_dimensions['A'].width = 25
            ws_details.column_dimensions['B'].width = 50
            ws_details.column_dimensions['C'].width = 70
            
            wb.save(output_file)
            
        elif export_format == "csv":
            logger.debug("Writing CSV output")
            headers = decoder.get_xlsx_headers()
            
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                writer.writerow(headers)
                
                # Write entries
                for entry in entries:
                    row = decoder.format_entry_for_xlsx(entry)
                    writer.writerow(row)
                
                # Add separator
                for _ in range(50):
                    writer.writerow([])
                
                # Write extraction details
                writer.writerow(["FENDER Extraction Report"])
                writer.writerow([])
                
                # System Information
                writer.writerow(["System Information"])
                writer.writerow(["Field", "Value"])
                for key, value in system_info.items():
                    if key != "decoder_hashes":
                        writer.writerow([key.replace("_", " ").title(), str(value)])
                
                writer.writerow([])
                
                # Decoder Hashes
                if "decoder_hashes" in system_info:
                    writer.writerow(["Decoder Integrity Verification"])
                    writer.writerow(["Decoder", "File Path", "SHA256 Hash", "File Size", "Last Modified"])
                    for decoder_name, hash_info in system_info["decoder_hashes"].items():
                        if "error" in hash_info:
                            writer.writerow([decoder_name, "Error", hash_info["error"], "", ""])
                        else:
                            writer.writerow([
                                decoder_name,
                                hash_info["file_path"],
                                hash_info["sha256_hash"],
                                hash_info["file_size"],
                                hash_info["last_modified"]
                            ])
                
                writer.writerow([])
                
                # Extraction Information
                writer.writerow(["Extraction Information"])
                writer.writerow(["Field", "Value"])
                
                # Input file details
                writer.writerow(["Input File Path", extraction_info["input_file"]["path"]])
                writer.writerow(["Input File Name", extraction_info["input_file"]["filename"]])
                writer.writerow(["Input File Size (MB)", extraction_info["input_file"]["size_mb"]])
                writer.writerow(["Input File SHA256", extraction_info["input_file"]["sha256_hash"]])
                
                # Output file details
                writer.writerow(["Output File Path", extraction_info["output_file"]["path"]])
                writer.writerow(["Output File Name", extraction_info["output_file"]["filename"]])
                
                # Extraction details
                writer.writerow(["Decoder Used", extraction_info["extraction_details"]["decoder_used"]])
                writer.writerow(["Entries Extracted", extraction_info["extraction_details"]["entries_extracted"]])
                writer.writerow(["Processing Time (seconds)", round(processing_time, 2)])
                    
        elif export_format == "json":
            logger.debug("Writing JSON output")
            json_data = {
                "metadata": {
                    "decoder": selected_decoder,
                    "extraction_timestamp": datetime.now().isoformat(),
                    "total_entries": len(entries)
                },
                "system_information": system_info,
                "extraction_information": extraction_info,
                "gps_entries": []
            }
            
            headers = decoder.get_xlsx_headers()
            
            for entry in entries:
                row = decoder.format_entry_for_xlsx(entry)
                entry_dict = {}
                
                for i, header in enumerate(headers):
                    if i < len(row):
                        entry_dict[header] = row[i]
                
                entry_dict.update({
                    "latitude": entry.latitude,
                    "longitude": entry.longitude,
                    "timestamp": entry.timestamp,
                    "extra_data": entry.extra_data
                })
                
                json_data["gps_entries"].append(entry_dict)
            
            with open(output_file, 'w', encoding='utf-8') as jsonfile:
                json.dump(json_data, jsonfile, indent=2, ensure_ascii=False, default=str)
        
        elif export_format == "geojson":
            logger.debug("Writing GeoJSON output")
            # Create GeoJSON with enhanced metadata
            features = []
            
            for i, entry in enumerate(entries):
                # Skip invalid coordinates
                if (entry.latitude == 0 and entry.longitude == 0) or \
                   not (-90 <= entry.latitude <= 90) or \
                   not (-180 <= entry.longitude <= 180):
                    continue
                
                # Create feature
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [entry.longitude, entry.latitude]
                    },
                    "properties": {
                        "id": i + 1,
                        "timestamp": entry.timestamp,
                        "latitude": entry.latitude,
                        "longitude": entry.longitude,
                    }
                }
                
                # Add extra data if available
                if entry.extra_data:
                    feature["properties"].update(entry.extra_data)
                
                features.append(feature)
            
            # Create GeoJSON structure with diagnostic data
            geojson = {
                "type": "FeatureCollection",
                "metadata": {
                    "decoder": selected_decoder,
                    "extraction_timestamp": datetime.now().isoformat(),
                    "total_features": len(features),
                    "coordinate_system": "WGS84",
                    "creator": f"FENDER v{FENDER_VERSION}",
                    "system_information": system_info,
                    "extraction_information": extraction_info
                },
                "features": features
            }
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(geojson, f, indent=2, ensure_ascii=False, default=str)

        elif export_format == "kml":
            logger.debug("Writing KML output")
            write_kml(entries, output_file, selected_decoder)
        
        print(f"\nSuccessfully extracted {len(entries)} GPS entries.")
        print(f"Results written to: {output_file}")
        logger.info(f"CLI processing complete. Output saved to: {output_file}")

def run_gui():
    """Run the GUI version"""
    logger.info("Starting FENDER in GUI mode")
    
    root = TkinterDnD.Tk()
    
    # Set icon
    try:
        icon_path = get_resource_path("car.ico")
        if os.path.exists(icon_path):
            root.iconbitmap(icon_path)
            logger.debug(f"Application icon set from: {icon_path}")
    except tk.TclError as e:
        logger.warning(f"Failed to set application icon: {e}")
    
    app = VehicleGPSDecoder(root)
    
    # Center window
    root.update_idletasks()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"+{position_x}+{position_y}")
    
    logger.info(f"GUI window centered at position: {position_x}, {position_y}")
    logger.info("Starting GUI main loop")
    
    root.mainloop()
    
    logger.info("GUI closed by user")

def main():
    logger.info("FENDER main() started")
    
    parser = argparse.ArgumentParser(
        description='Vehicle GPS Decoder - Extract GPS data from vehicle telematics binary files'
    )
    parser.add_argument('--cli', action='store_true', help='Run in command line interface mode')
    
    args = parser.parse_args()
    logger.info(f"Command line arguments: {sys.argv[1:]}")
    
    if args.cli:
        logger.info("Running in CLI mode")
        run_cli()
    else:
        logger.info("Running in GUI mode")
        run_gui()
    
    logger.info("FENDER main() completed")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.critical(f"Unhandled exception in main: {e}", exc_info=True)
        raise