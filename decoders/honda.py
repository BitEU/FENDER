#!/usr/bin/env python3
"""
Honda CRM Database Extractor for Windows with Excel Export

This script extracts the crm.db file from Honda Android system images,
opens the eco_logs table, and exports specific columns to Excel.

Usage:
    python honda_crm_extractor_enhanced.py <image_file> [output_directory]

Requirements:
    pip install pytsk3 openpyxl
    
For Windows: Download pytsk3 wheel from:
https://github.com/py4n6/pytsk/releases
"""

import os
import sys
import struct
import argparse
import logging
import sqlite3
from pathlib import Path
from typing import Optional, Tuple, BinaryIO

try:
    import pytsk3
    TSK_AVAILABLE = True
except ImportError:
    TSK_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class WindowsAndroidImageExtractor:
    """Extract files from Android system images on Windows."""
    
    def __init__(self, image_path: str):
        self.image_path = Path(image_path)
        self.image_size = self.image_path.stat().st_size
        self.temp_files = []  # Track temporary files for cleanup
        logger.info(f"Image size: {self.image_size / (1024**3):.2f} GB")
    
    def find_partition_by_name(self, partition_name: str = "userdata") -> Optional[Tuple[int, int]]:
        """
        Find partition offset and size by scanning for GPT or common Android patterns.
        Returns (offset, size) tuple or None if not found.
        """
        with open(self.image_path, 'rb') as f:
            # Try to find GPT header first
            gpt_offset, gpt_size = self._find_gpt_partition(f, partition_name)
            if gpt_offset:
                return gpt_offset, gpt_size
            
            # Try to find ext4 signature directly
            ext4_offset, ext4_size = self._find_ext4_partition(f)
            if ext4_offset:
                return ext4_offset, ext4_size
        
        return None
    
    def _find_gpt_partition(self, f: BinaryIO, partition_name: str) -> Tuple[Optional[int], Optional[int]]:
        """Find partition using GPT (GUID Partition Table)."""
        try:
            # GPT header is at LBA 1 (sector size 512)
            f.seek(512)
            gpt_header = f.read(92)
            
            if len(gpt_header) < 92 or gpt_header[:8] != b'EFI PART':
                logger.debug("No valid GPT header found")
                return None, None
            
            logger.info("Found GPT header, parsing partitions...")
            
            # Parse GPT header
            partition_entries_lba, = struct.unpack('<Q', gpt_header[72:80])
            num_partitions, = struct.unpack('<I', gpt_header[80:84])
            partition_entry_size, = struct.unpack('<I', gpt_header[84:88])
            
            # Read partition entries
            f.seek(partition_entries_lba * 512)
            
            for i in range(min(num_partitions, 128)):  # Reasonable limit
                entry = f.read(partition_entry_size)
                if len(entry) < 128:
                    break
                
                # Check if partition exists (non-zero GUID)
                if entry[:16] == b'\x00' * 16:
                    continue
                
                # Extract partition name (UTF-16LE, 72 bytes max)
                name_bytes = entry[56:128]
                try:
                    name = name_bytes.decode('utf-16le').rstrip('\x00')
                except:
                    continue
                
                logger.debug(f"Found partition: {name}")
                
                if partition_name.lower() in name.lower():
                    start_lba, = struct.unpack('<Q', entry[32:40])
                    end_lba, = struct.unpack('<Q', entry[40:48])
                    
                    offset = start_lba * 512
                    size = (end_lba - start_lba + 1) * 512
                    
                    logger.info(f"Found {name} partition at offset {offset} (size: {size / (1024**3):.2f} GB)")
                    return offset, size
            
        except Exception as e:
            logger.debug(f"GPT parsing failed: {e}")
        
        return None, None
    
    def _find_ext4_partition(self, f: BinaryIO) -> Tuple[Optional[int], Optional[int]]:
        """Find ext4 partition by scanning for superblock signature."""
        logger.info("Scanning for ext4 superblock signatures...")
        
        # ext4 superblock is at offset 1024 from partition start
        # Scan in larger chunks for speed, but limit search area
        chunk_size = 4 * 1024 * 1024  # 4MB chunks
        max_search = min(self.image_size, 500 * 1024 * 1024)  # Search first 500MB
        
        for offset in range(0, max_search, chunk_size):
            if offset % (50 * 1024 * 1024) == 0:  # Progress every 50MB
                logger.info(f"Scanning at {offset / (1024**2):.0f} MB...")
            
            f.seek(offset)
            chunk = f.read(min(chunk_size, max_search - offset))
            
            # Look for ext4 magic number (0xEF53) at offset 1024 + 56
            for i in range(len(chunk) - 1080):
                if chunk[i + 1024 + 56:i + 1024 + 58] == b'\x53\xEF':
                    partition_offset = offset + i
                    
                    # Validate this is a real ext4 superblock
                    f.seek(partition_offset + 1024)
                    superblock = f.read(1024)
                    
                    if len(superblock) >= 1024:
                        # Get block count and block size
                        block_count, = struct.unpack('<I', superblock[4:8])
                        log_block_size, = struct.unpack('<I', superblock[24:28])
                        block_size = 1024 << log_block_size
                        
                        # Sanity check
                        if block_size in [1024, 2048, 4096] and block_count > 1000:
                            partition_size = block_count * block_size
                            
                            logger.info(f"Found ext4 partition at offset {partition_offset} (size: {partition_size / (1024**3):.2f} GB)")
                            logger.info(f"Block size: {block_size}, Block count: {block_count}")
                            return partition_offset, partition_size
        
        logger.warning("No ext4 partition found")
        return None, None
    
    def extract_and_process_crm_database(self, output_dir: str = "extracted") -> bool:
        """Extract the Honda CRM database and process eco_logs to Excel."""
        # Find userdata partition
        partition_info = self.find_partition_by_name("userdata")
        if not partition_info:
            logger.error("Could not find userdata partition")
            logger.info("Try running with -v for verbose output to debug partition detection")
            return False
        
        offset, size = partition_info
        
        # Create output directory
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        if not TSK_AVAILABLE:
            logger.error("pytsk3 is required for Windows extraction!")
            logger.error("Install with: pip install pytsk3")
            logger.error("Or download from: https://github.com/py4n6/pytsk/releases")
            return False
        
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl is required for Excel export!")
            logger.error("Install with: pip install openpyxl")
            return False
        
        try:
            # Extract the database
            crm_db_path = self._extract_with_tsk(offset, size, output_path)
            if not crm_db_path:
                return False
            
            # Process the database and create Excel file
            excel_path = self._process_eco_logs_to_excel(crm_db_path, output_path)
            if not excel_path:
                return False
            
            # Clean up temporary files
            self._cleanup_temp_files()
            
            logger.info("=" * 60)
            logger.info("SUCCESS! Honda CRM eco_logs data exported to Excel!")
            logger.info(f"Excel file created: {excel_path}")
            logger.info("All temporary files have been cleaned up.")
            logger.info("=" * 60)
            
            return True
            
        except Exception as e:
            logger.error(f"Processing failed: {e}")
            self._cleanup_temp_files()
            return False
    
    def _extract_with_tsk(self, offset: int, size: int, output_path: Path) -> Optional[Path]:
        """Extract using The Sleuth Kit (pytsk3)."""
        logger.info("Using pytsk3 for ext4 extraction")
        
        try:
            # Create a subset image for the partition
            subset_path = output_path / "userdata_partition.img"
            self.temp_files.append(subset_path)  # Track for cleanup
            
            logger.info("Extracting userdata partition to temporary file...")
            self._extract_partition_to_file(offset, size, subset_path)
            
            # Open with TSK
            logger.info("Opening partition with TSK...")
            img = pytsk3.Img_Info(str(subset_path))
            fs = pytsk3.FS_Info(img)
            
            # Try multiple possible paths for Honda CRM database
            crm_db_path = self._try_extract_crm_paths(fs, output_path)
            
            return crm_db_path
                
        except Exception as e:
            logger.error(f"TSK extraction failed: {e}")
            logger.error("This might be due to:")
            logger.error("1. Corrupted or unsupported filesystem")
            logger.error("2. Encrypted userdata partition") 
            logger.error("3. Non-standard ext4 features")
            return None
    
    def _extract_partition_to_file(self, offset: int, size: int, output_file: Path):
        """Extract partition data to a temporary file."""
        with open(self.image_path, 'rb') as src:
            src.seek(offset)
            with open(output_file, 'wb') as dst:
                remaining = size
                chunk_size = 8 * 1024 * 1024  # 8MB chunks for speed
                
                while remaining > 0:
                    read_size = min(chunk_size, remaining)
                    chunk = src.read(read_size)
                    if not chunk:
                        break
                    dst.write(chunk)
                    remaining -= len(chunk)
                    
                    # Progress indicator
                    progress = ((size - remaining) / size) * 100
                    if int(progress) % 10 == 0:
                        logger.info(f"Extraction progress: {progress:.0f}%")
    
    def _try_extract_crm_paths(self, fs, output_path: Path) -> Optional[Path]:
        """Try extracting CRM database from multiple possible paths."""
        
        # Common paths where Honda CRM database might be located
        search_paths = [
            "/data/com.honda.telematics.core/databases/crm.db",
            "/data/data/com.honda.telematics.core/databases/crm.db", 
            "data/com.honda.telematics.core/databases/crm.db",
            "data/data/com.honda.telematics.core/databases/crm.db",
            "/userdata/data/com.honda.telematics.core/databases/crm.db",
            "userdata/data/com.honda.telematics.core/databases/crm.db",
        ]
        
        for search_path in search_paths:
            try:
                logger.info(f"Trying path: {search_path}")
                file_obj = fs.open(search_path)
                
                crm_db_path = output_path / "crm.db"
                logger.info(f"Extracting crm.db to {crm_db_path}")
                
                # Read and write the file
                file_size = file_obj.info.meta.size
                data = file_obj.read_random(0, file_size)
                
                with open(crm_db_path, 'wb') as output_file:
                    output_file.write(data)
                
                file_size = crm_db_path.stat().st_size
                logger.info(f"Successfully extracted crm.db ({file_size / 1024:.1f} KB)")
                return crm_db_path
                
            except Exception as e:
                logger.debug(f"Path {search_path} failed: {e}")
                continue
        
        # If direct paths fail, try recursive search
        logger.info("Direct paths failed, starting recursive search...")
        return self._recursive_search_crm(fs, output_path)
    
    def _recursive_search_crm(self, fs, output_path: Path) -> Optional[Path]:
        """Recursively search for crm.db files."""
        try:
            if self._search_directory(fs, fs.open_dir("/"), output_path, "/", 0):
                # Return the first crm.db found
                crm_files = list(output_path.glob("crm*.db"))
                return crm_files[0] if crm_files else None
            return None
        except Exception as e:
            logger.error(f"Recursive search failed: {e}")
            return None
    
    def _search_directory(self, fs, directory, output_path: Path, current_path: str, depth: int) -> bool:
        """Search directory recursively with depth limit."""
        if depth > 10:  # Prevent infinite recursion
            return False
            
        try:
            for entry in directory:
                entry_name = entry.info.name.name.decode('utf-8', errors='ignore')
                
                if entry_name in ['.', '..']:
                    continue
                
                full_path = f"{current_path}{entry_name}" if current_path.endswith('/') else f"{current_path}/{entry_name}"
                
                # Check if it's a regular file named crm.db
                if (entry.info.meta and 
                    entry.info.meta.type == pytsk3.TSK_FS_META_TYPE_REG and 
                    entry_name.lower() == "crm.db"):
                    
                    logger.info(f"Found crm.db at: {full_path}")
                    
                    try:
                        file_obj = fs.open(full_path)
                        crm_db_path = output_path / f"crm_{depth}_{len(list(output_path.glob('crm*.db')))}.db"
                        
                        # Read file data and write to output
                        file_size = file_obj.info.meta.size
                        data = file_obj.read_random(0, file_size)
                        
                        with open(crm_db_path, 'wb') as output_file:
                            output_file.write(data)
                        
                        file_size = crm_db_path.stat().st_size
                        logger.info(f"Extracted to {crm_db_path} ({file_size / 1024:.1f} KB)")
                        return True
                    except Exception as e:
                        logger.debug(f"Failed to extract {full_path}: {e}")
                
                # Recurse into directories that might contain Honda data
                elif (entry.info.meta and 
                      entry.info.meta.type == pytsk3.TSK_FS_META_TYPE_DIR and
                      any(keyword in entry_name.lower() for keyword in ['honda', 'telematics', 'data', 'app'])):
                    
                    try:
                        sub_dir = fs.open_dir(full_path)
                        if self._search_directory(fs, sub_dir, output_path, full_path, depth + 1):
                            return True
                    except Exception as e:
                        logger.debug(f"Cannot access directory {full_path}: {e}")
                        continue
        
        except Exception as e:
            logger.debug(f"Directory search error at {current_path}: {e}")
        
        return False
    
    def _process_eco_logs_to_excel(self, crm_db_path: Path, output_path: Path) -> Optional[Path]:
        """Process the eco_logs table and export to Excel."""
        logger.info("Processing eco_logs table from CRM database...")
        
        try:
            # Connect to the SQLite database
            conn = sqlite3.connect(str(crm_db_path))
            cursor = conn.cursor()
            
            # Check if eco_logs table exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='eco_logs';")
            if not cursor.fetchone():
                logger.error("eco_logs table not found in the database")
                
                # List available tables for debugging
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                tables = cursor.fetchall()
                logger.info(f"Available tables: {[table[0] for table in tables]}")
                conn.close()
                return None
            
            # Check if the required columns exist
            cursor.execute("PRAGMA table_info(eco_logs);")
            columns = [row[1] for row in cursor.fetchall()]
            logger.info(f"Available columns in eco_logs: {columns}")
            
            required_columns = [
                'start_pos_time', 'start_pos_lat', 'start_pos_lon',
                'finish_pos_time', 'finish_pos_lat', 'finish_pos_lon'
            ]
            
            # Check which columns are available
            available_columns = [col for col in required_columns if col in columns]
            missing_columns = [col for col in required_columns if col not in columns]
            
            if missing_columns:
                logger.warning(f"Missing columns: {missing_columns}")
                logger.info(f"Will export available columns: {available_columns}")
            
            if not available_columns:
                logger.error("None of the required columns found in eco_logs table")
                conn.close()
                return None
            
            # Query the available columns
            columns_str = ', '.join(available_columns)
            query = f"SELECT {columns_str} FROM eco_logs"
            logger.info(f"Executing query: {query}")
            
            cursor.execute(query)
            rows = cursor.fetchall()
            
            if not rows:
                logger.warning("No data found in eco_logs table")
                conn.close()
                return None
            
            logger.info(f"Found {len(rows)} records in eco_logs table")
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "eco_logs_data"
            
            # Write headers
            for col_idx, col_name in enumerate(available_columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel file
            excel_path = output_path / "honda_eco_logs.xlsx"
            wb.save(str(excel_path))
            
            file_size = excel_path.stat().st_size
            logger.info(f"Excel file created: {excel_path} ({file_size / 1024:.1f} KB)")
            logger.info(f"Exported {len(rows)} records with {len(available_columns)} columns")
            
            conn.close()
            return excel_path
            
        except Exception as e:
            logger.error(f"Failed to process eco_logs table: {e}")
            try:
                conn.close()
            except:
                pass
            return None
    
    def _cleanup_temp_files(self):
        """Clean up all temporary files."""
        logger.info("Cleaning up temporary files...")
        
        for temp_file in self.temp_files:
            try:
                if temp_file.exists():
                    temp_file.unlink()
                    logger.debug(f"Deleted temporary file: {temp_file}")
            except Exception as e:
                logger.warning(f"Failed to delete {temp_file}: {e}")
        
        self.temp_files.clear()


def main():
    parser = argparse.ArgumentParser(description="Extract Honda CRM eco_logs data to Excel (Windows)")
    parser.add_argument("image_file", help="Path to the Android system image file")
    parser.add_argument("-o", "--output", default="extracted", help="Output directory (default: extracted)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Enable verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    if not Path(args.image_file).exists():
        logger.error(f"Image file not found: {args.image_file}")
        sys.exit(1)
    
    if not TSK_AVAILABLE:
        logger.error("pytsk3 is required but not installed!")
        logger.error("Install with: pip install pytsk3")
        logger.error("For Windows, you may need to download a wheel from:")
        logger.error("https://github.com/py4n6/pytsk/releases")
        sys.exit(1)
    
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is required but not installed!")
        logger.error("Install with: pip install openpyxl")
        sys.exit(1)
    
    extractor = WindowsAndroidImageExtractor(args.image_file)
    
    logger.info("Starting Honda CRM eco_logs extraction and Excel export...")
    success = extractor.extract_and_process_crm_database(args.output)
    
    if success:
        sys.exit(0)
    else:
        logger.error("=" * 50)
        logger.error("EXTRACTION AND PROCESSING FAILED!")
        logger.error("Try running with -v for more detailed debugging info")
        logger.error("=" * 50)
        sys.exit(1)


if __name__ == "__main__":
    main()