import os
import re
import struct
import json
from datetime import datetime, timezone
from typing import List, Tuple, Optional, Any
from src.core.base_decoder import BaseDecoder, GPSEntry
import logging
import time

# Setup logger for this module
logger = logging.getLogger(__name__)

class DensoDecoder(BaseDecoder):
    """
    Denso Vehicle Decoder
    Extracts GPS data from Denso vehicle telematics binary files
    """
    
    def __init__(self):
        super().__init__()
          # GPS data extraction patterns
        self.gps_patterns = {
            'Navigation.Location': {
                'pattern': rb'\{"timestamp":.*?,"tag":"Navigation\.Location"',
                'has_coordinates': True,
                'has_speed': True
            },
            'Frame.VehicleSpeed': {
                'pattern': rb'\{"timestamp":.*?,"tag":"Frame\.VehicleSpeed"',
                'has_coordinates': False,
                'has_speed': True
            },
            'Phone.BluetoothConnection': {
                'pattern': rb'\{"timestamp":.*?,"tag":"Phone\.BluetoothConnection"',
                'has_coordinates': False,
                'has_bluetooth': True
            }        }
        
        self._logger.info("DensoDecoder initialized")
        self._logger.debug(f"Configured {len(self.gps_patterns)} GPS pattern types")
    
    def get_name(self) -> str:
        return "Acura Denso DNNS087"
    
    def get_supported_extensions(self) -> List[str]:
        extensions = ['.bin', '.001', '.CE0']
        self._logger.debug(f"Supported extensions: {extensions}")
        return extensions
    
    def get_dropzone_text(self) -> str:
        return "Drop your Denso telematics binary file\nhere or click to browse"

    def get_xlsx_headers(self) -> List[str]:
        headers = [
            'Unix Epoch',
            'Timestamp (UTC)',
            'Event Type',
            'Latitude',
            'Longitude',
            'Accuracy',
            'Speed (KMH)',
            'Bearing',
            'Vehicle Speed (KMH)',
            'Bluetooth Device',
            'Bluetooth State'
        ]
        self._logger.debug(f"XLSX headers: {len(headers)} columns")
        return headers

    def get_location_headers(self) -> List[str]:
        """Headers specific to location data"""
        return [
            'Unix Epoch',
            'Timestamp (UTC)',
            'Event Type',
            'Latitude',
            'Longitude',
            'Accuracy',
            'Speed (KMH)',
            'Bearing',
            'Fix Time'
        ]

    def get_speed_headers(self) -> List[str]:
        """Headers specific to speed data"""
        return [
            'Unix Epoch',
            'Timestamp (UTC)',
            'Event Type',
            'Vehicle Speed (KMH)'
        ]

    def get_bluetooth_headers(self) -> List[str]:
        """Headers specific to bluetooth data"""
        return [
            'Unix Epoch',
            'Timestamp (UTC)',
            'Event Type',
            'Bluetooth Device',
            'Bluetooth State'
        ]

    def categorize_entries_by_type(self, entries: List[GPSEntry]) -> dict:
        """Categorize entries by their event type"""
        categorized = {
            'location': [],
            'speed': [],
            'bluetooth': []
        }
        
        for entry in entries:
            event_type = entry.extra_data.get('event_type', '')
            if event_type == 'Navigation.Location':
                categorized['location'].append(entry)
            elif event_type == 'Frame.VehicleSpeed':
                categorized['speed'].append(entry)
            elif event_type == 'Phone.BluetoothConnection':
                categorized['bluetooth'].append(entry)
        
        self._logger.info(f"Categorized entries: Location={len(categorized['location'])}, "
                         f"Speed={len(categorized['speed'])}, Bluetooth={len(categorized['bluetooth'])}")
        return categorized

    def format_location_entry_for_xlsx(self, entry: GPSEntry) -> List[Any]:
        """Format a location entry for XLSX export"""
        return [
            entry.extra_data.get('unix_epoch', ''),
            entry.timestamp if entry.timestamp else '',
            entry.extra_data.get('event_type', ''),
            entry.latitude if entry.latitude != 0 else '',
            entry.longitude if entry.longitude != 0 else '',
            entry.extra_data.get('accuracy', ''),
            entry.extra_data.get('speed', ''),
            entry.extra_data.get('bearing', ''),
            entry.extra_data.get('fix_time', '')
        ]

    def format_speed_entry_for_xlsx(self, entry: GPSEntry) -> List[Any]:
        """Format a speed entry for XLSX export"""
        return [
            entry.extra_data.get('unix_epoch', ''),
            entry.timestamp if entry.timestamp else '',
            entry.extra_data.get('event_type', ''),
            entry.extra_data.get('vehicle_speed_kmh', '')
        ]

    def format_bluetooth_entry_for_xlsx(self, entry: GPSEntry) -> List[Any]:
        """Format a bluetooth entry for XLSX export"""
        return [
            entry.extra_data.get('unix_epoch', ''),
            entry.timestamp if entry.timestamp else '',
            entry.extra_data.get('event_type', ''),
            entry.extra_data.get('bluetooth_device', ''),
            entry.extra_data.get('bluetooth_state', '')
        ]

    def format_entry_for_xlsx(self, entry: GPSEntry) -> List[Any]:
        """Format a GPSEntry into a row for the XLSX file"""
        self._logger.debug(f"Formatting entry for XLSX: lat={entry.latitude}, lon={entry.longitude}")
        
        row = [
            entry.extra_data.get('unix_epoch', ''),
            entry.timestamp if entry.timestamp else '',
            entry.extra_data.get('event_type', ''),
            entry.latitude if entry.latitude != 0 else '',
            entry.longitude if entry.longitude != 0 else '',
            entry.extra_data.get('accuracy', ''),
            entry.extra_data.get('speed', ''),
            entry.extra_data.get('bearing', ''),
            entry.extra_data.get('vehicle_speed_kmh', ''),
            entry.extra_data.get('bluetooth_device', ''),
            entry.extra_data.get('bluetooth_state', '')
        ]
        
        return row
    
    def extract_gps_data(self, file_path: str, progress_callback=None, stop_event=None) -> Tuple[List[GPSEntry], Optional[str]]:
        """
        Extract GPS data from Denso binary file
        
        Args:
            file_path: Path to the input file
            progress_callback: Optional callback for progress updates
            stop_event: Optional threading.Event to signal stop processing
            
        Returns:
            Tuple of (GPS entries list, error message or None)
        """
        start_time = time.time()
        self._log_extraction_start(file_path)
        
        entries = []
        
        try:
            # Check file size
            file_size = os.path.getsize(file_path)
            self._logger.info(f"Processing Denso file: {file_path} (Size: {file_size/1024/1024:.2f} MB)")
            
            if progress_callback:
                progress_callback("Reading binary file...", 10)
                self._log_progress("Reading binary file", 10)
            
            # Check for stop signal
            if stop_event and stop_event.is_set():
                self._logger.warning("Processing stopped by user before file read")
                return [], "Processing stopped by user."
            
            # Read file
            self._logger.debug("Opening file for binary read")
            with open(file_path, 'rb') as f:
                data = f.read()
            
            self._logger.info(f"Successfully read {len(data)} bytes from file")
            
            if progress_callback:
                progress_callback("Searching for GPS data patterns...", 30)
                self._log_progress("Searching for GPS data patterns", 30)
            
            # Check for stop signal
            if stop_event and stop_event.is_set():
                self._logger.warning("Processing stopped by user before pattern search")
                return [], "Processing stopped by user."
            
            # Process data using the boundary-to-boundary parsing strategy
            all_records = self._extract_records(data, progress_callback, stop_event)
              # Convert records to GPSEntry objects
            entries = self._convert_to_gps_entries(all_records)
            
            elapsed_time = time.time() - start_time
            
            if progress_callback:
                progress_callback("Processing complete!", 100)
                self._log_progress("Processing complete", 100)
            
            self._log_extraction_complete(len(entries), elapsed_time)
            return entries, None
            
        except FileNotFoundError:
            error_msg = f"File not found: {file_path}"
            self._log_extraction_error(error_msg)
            return [], error_msg
        except PermissionError:
            error_msg = f"Permission denied accessing file: {file_path}"
            self._log_extraction_error(error_msg)
            return [], error_msg
        except Exception as e:
            error_msg = f"Error processing file: {str(e)}"
            self._logger.error(f"Unexpected error during extraction: {e}", exc_info=True)
            self._log_extraction_error(error_msg)
            return [], error_msg
    
    def _extract_records(self, data: bytes, progress_callback=None, stop_event=None) -> dict:
        """Extract all records from binary data"""
        self._logger.info("Starting record extraction from binary data")
          # Results dictionary for different tag types
        results = {
            'Navigation.Location': [],
            'Frame.VehicleSpeed': [],
            'Phone.BluetoothConnection': []
        }
          # Combined regex to match all desired tags
        start_regex = re.compile(
            rb'\{"timestamp":.*?,"tag":"(?:Navigation\.Location|Frame\.VehicleSpeed|Phone\.BluetoothConnection)"'
        )
        next_record_marker = b',{"timestamp":'
        
        matches = list(start_regex.finditer(data))
        total_matches = len(matches)
        self._logger.info(f"Found {total_matches} potential records")
        
        if progress_callback:
            progress_callback(f"Found {total_matches} records to process...", 40)
        
        for i, match in enumerate(matches):
            # Check for stop signal periodically
            if stop_event and stop_event.is_set() and i % 100 == 0:
                self._logger.warning(f"Processing stopped by user at record {i}/{total_matches}")
                break
            
            obj_start_pos = match.start()
            obj_end_pos = data.find(next_record_marker, obj_start_pos + 1)
            
            if obj_end_pos == -1:
                slice_to_parse = data[obj_start_pos : obj_start_pos + 4096]
            else:
                slice_to_parse = data[obj_start_pos : obj_end_pos]
            
            try:
                # Parse JSON - decode bytes to string first
                # Decode bytes to string, handling potential encoding issues
                try:
                    json_str = slice_to_parse.decode('utf-8')
                except UnicodeDecodeError:
                    # If UTF-8 fails, try other encodings or skip this record
                    try:
                        json_str = slice_to_parse.decode('latin-1')
                    except UnicodeDecodeError:
                        self._logger.debug(f"Failed to decode bytes at position {obj_start_pos}")
                        continue
                
                record_data = json.loads(json_str)
                tag = record_data.get('tag', '')
                
                if tag in results:
                    results[tag].append(record_data)
                
            except json.JSONDecodeError:
                self._logger.debug(f"Failed to parse JSON at position {obj_start_pos}")
                continue
            
            # Update progress
            if progress_callback and i % 1000 == 0:
                progress = 40 + (40 * i // total_matches)
                progress_callback(f"Processing record {i}/{total_matches}...", progress)
                if i % 10000 == 0:
                    self._log_progress(f"Processing records ({i}/{total_matches})", progress)
        
        # Log summary
        for tag_type, records in results.items():
            self._logger.info(f"{tag_type}: {len(records)} records extracted")
        
        return results
    
    def _convert_to_gps_entries(self, all_records: dict) -> List[GPSEntry]:
        """Convert extracted records to GPSEntry objects"""
        self._logger.info("Converting records to GPS entries")
        entries = []
        
        # Process Navigation.Location records (primary GPS data)
        for record in all_records.get('Navigation.Location', []):
            entry = self._process_navigation_location(record)
            if entry:
                entries.append(entry)
          # Process other record types and add as entries with limited location data
        # These might not have coordinates but contain other valuable data
        
        # Vehicle speed data
        for record in all_records.get('Frame.VehicleSpeed', []):
            entry = self._process_vehicle_speed(record)
            if entry:
                entries.append(entry)
        
        # Bluetooth connection events
        for record in all_records.get('Phone.BluetoothConnection', []):
            entry = self._process_bluetooth(record)
            if entry:
                entries.append(entry)
        
        # Sort entries by timestamp
        entries.sort(key=lambda x: x.extra_data.get('unix_epoch', 0))
        
        self._logger.info(f"Created {len(entries)} GPS entries")
        return entries
    
    def _process_navigation_location(self, record: dict) -> Optional[GPSEntry]:
        """Process Navigation.Location record"""
        try:
            timestamp = record.get('timestamp', '')
            unix_epoch = self._convert_timestamp_to_unix(timestamp)
            
            value = record.get('value', {})
            coordinate = value.get('coordinate', {})
            velocity = value.get('velocity', {})
            
            latitude = coordinate.get('latitude')
            longitude = coordinate.get('longitude')
            
            # Validate coordinates
            if not self._is_valid_coordinate(latitude, longitude):
                return None
            
            # Get additional fix time if available
            fix_time_str = value.get('fixTime')
            fix_time_unix = self._convert_timestamp_to_unix(fix_time_str) if fix_time_str else None
            
            extra_data = {
                'unix_epoch': unix_epoch,
                'event_type': 'Navigation.Location',
                'accuracy': value.get('accuracy', ''),
                'speed': velocity.get('speed', ''),
                'bearing': velocity.get('bearing', ''),
                'fix_time': fix_time_str if fix_time_str else ''
            }
            
            return GPSEntry(
                latitude=latitude,
                longitude=longitude,
                timestamp=timestamp,
                extra_data=extra_data
            )
            
        except Exception as e:
            self._logger.debug(f"Error processing navigation location: {e}")
            return None
    def _process_vehicle_speed(self, record: dict) -> Optional[GPSEntry]:
        """Process Frame.VehicleSpeed record"""
        try:
            timestamp = record.get('timestamp', '')
            unix_epoch = self._convert_timestamp_to_unix(timestamp)
            
            value = record.get('value', {})
            
            extra_data = {
                'unix_epoch': unix_epoch,
                'event_type': 'Frame.VehicleSpeed',
                'vehicle_speed_kmh': value.get('kilometersPerHour', '')
            }
            
            # Speed data doesn't have coordinates, use 0,0
            return GPSEntry(
                latitude=0,
                longitude=0,
                timestamp=timestamp,
                extra_data=extra_data
            )
            
        except Exception as e:
            self._logger.debug(f"Error processing vehicle speed: {e}")
            return None
    
    def _process_bluetooth(self, record: dict) -> Optional[GPSEntry]:
        """Process Phone.BluetoothConnection record"""
        try:
            timestamp = record.get('timestamp', '')
            unix_epoch = self._convert_timestamp_to_unix(timestamp)
            
            value = record.get('value', {})
            
            device_name = value.get('deviceName', '')
            device_id = value.get('deviceId', '')
            device_address = value.get('deviceAddress', '')
            
            # Create device identifier
            bluetooth_device = device_name or device_id or device_address or 'Unknown'
            
            extra_data = {
                'unix_epoch': unix_epoch,
                'event_type': 'Phone.BluetoothConnection',
                'bluetooth_device': bluetooth_device,
                'bluetooth_state': value.get('state', '')
            }
            
            # Bluetooth data doesn't have coordinates, use 0,0
            return GPSEntry(
                latitude=0,
                longitude=0,
                timestamp=timestamp,
                extra_data=extra_data
            )
            
        except Exception as e:
            self._logger.debug(f"Error processing bluetooth connection: {e}")
            return None
    
    def _convert_timestamp_to_unix(self, timestamp_str: str) -> float:
        """Convert ISO timestamp to Unix epoch"""
        if timestamp_str and isinstance(timestamp_str, str):
            try:
                # Replace 'Z' with timezone info for robust parsing
                dt_object = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                return dt_object.timestamp()
            except ValueError as e:
                self._logger.debug(f"Failed to convert timestamp '{timestamp_str}': {e}")
                return 0
        return 0
    
    def _is_valid_coordinate(self, lat: float, lon: float) -> bool:
        """Check if coordinates are valid GPS values"""
        if lat is None or lon is None:
            return False
        
        # Basic range check
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            return False
          # Check for null island
        if lat == 0 and lon == 0:
            return False
        
        return True
    
    def export_to_excel_with_separate_sheets(self, entries: List[GPSEntry], output_path: str, 
                                           decoder_name: str, system_info: dict, 
                                           extraction_info: dict, examiner_name: str = None, 
                                           case_number: str = None):
        """Export data to Excel with separate worksheets for different data types"""
        self._logger.info(f"Writing Excel report with separate sheets to: {output_path}")
        
        from openpyxl import Workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Categorize entries
        categorized = self.categorize_entries_by_type(entries)
        
        # Create Location Data worksheet
        if categorized['location']:
            ws_location = wb.create_sheet("Location Data")
            headers = self.get_location_headers()
            ws_location.append(headers)
            
            for entry in categorized['location']:
                row = self.format_location_entry_for_xlsx(entry)
                ws_location.append(row)
            self._logger.info(f"Added {len(categorized['location'])} location entries to worksheet")
        
        # Create Speed Data worksheet
        if categorized['speed']:
            ws_speed = wb.create_sheet("Speed Data")
            headers = self.get_speed_headers()
            ws_speed.append(headers)
            
            for entry in categorized['speed']:
                row = self.format_speed_entry_for_xlsx(entry)
                ws_speed.append(row)
            self._logger.info(f"Added {len(categorized['speed'])} speed entries to worksheet")
        
        # Create Bluetooth Data worksheet
        if categorized['bluetooth']:
            ws_bluetooth = wb.create_sheet("Bluetooth Data")
            headers = self.get_bluetooth_headers()
            ws_bluetooth.append(headers)
            
            for entry in categorized['bluetooth']:
                row = self.format_bluetooth_entry_for_xlsx(entry)
                ws_bluetooth.append(row)
            self._logger.info(f"Added {len(categorized['bluetooth'])} bluetooth entries to worksheet")
        
        # Create Extraction Details worksheet (same as original)
        ws_details = wb.create_sheet("Extraction Details")
        
        # Write extraction details
        ws_details.append(["FENDER Extraction Report"])
        ws_details.append([])
        
        # Case Information (if provided)
        if examiner_name or case_number:
            ws_details.append(["Case Information"])
            ws_details.append(["Field", "Value"])
            if examiner_name:
                ws_details.append(["Examiner Name", examiner_name])
            if case_number:
                ws_details.append(["Case Number", case_number])
            ws_details.append([])
        
        # System Information
        ws_details.append(["System Information"])
        ws_details.append(["Field", "Value"])
        for key, value in system_info.items():
            if key != "decoder_hashes":
                ws_details.append([key.replace("_", " ").title(), str(value)])
        
        ws_details.append([])
        
        # Decoder Hashes
        if "decoder_hashes" in system_info:
            ws_details.append(["Decoder Integrity Verification"])
            ws_details.append(["Decoder", "File Path", "SHA256 Hash", "File Size", "Last Modified"])
            for decoder_name_hash, hash_info in system_info["decoder_hashes"].items():
                if "error" in hash_info:
                    ws_details.append([decoder_name_hash, "Error", hash_info["error"], "", ""])
                else:
                    ws_details.append([
                        decoder_name_hash,
                        hash_info["file_path"],
                        hash_info["sha256_hash"],
                        hash_info["file_size"],
                        hash_info["last_modified"]
                    ])
        
        ws_details.append([])
        
        # Extraction Information
        ws_details.append(["Extraction Information"])
        ws_details.append(["Field", "Value"])
        
        # Input file details
        ws_details.append(["Input File Path", extraction_info["input_file"]["path"]])
        ws_details.append(["Input File Name", extraction_info["input_file"]["filename"]])
        ws_details.append(["Input File Size (MB)", extraction_info["input_file"]["size_mb"]])
        ws_details.append(["Input File SHA256", extraction_info["input_file"]["sha256_hash"]])
        
        # Output file details
        ws_details.append(["Output File Path", extraction_info["output_file"]["path"]])
        ws_details.append(["Output File Name", extraction_info["output_file"]["filename"]])
        
        # Extraction details
        ws_details.append(["Decoder Used", extraction_info["extraction_details"]["decoder_used"]])
        ws_details.append(["Entries Extracted", extraction_info["extraction_details"]["entries_extracted"]])
        ws_details.append(["Processing Time (seconds)", extraction_info["extraction_details"]["processing_time_seconds"]])
        
        # Format the details worksheet
        ws_details.column_dimensions['A'].width = 25
        ws_details.column_dimensions['B'].width = 50
        ws_details.column_dimensions['C'].width = 70
        wb.save(output_path)
        self._logger.info(f"Excel report with separate sheets written successfully: {output_path}")

    def export_to_json_with_separate_sections(self, entries: List[GPSEntry], output_path: str,
                                            decoder_name: str, system_info: dict,
                                            extraction_info: dict, examiner_name: str = None,
                                            case_number: str = None):
        """Export data to JSON with separate sections for different data types"""
        self._logger.info(f"Writing JSON report with separate sections to: {output_path}")
        
        # Categorize entries
        categorized = self.categorize_entries_by_type(entries)
        
        json_data = {
            "metadata": {
                "decoder": decoder_name,
                "extraction_timestamp": datetime.now().isoformat(),
                "total_entries": len(entries),
                "location_entries": len(categorized['location']),
                "speed_entries": len(categorized['speed']),
                "bluetooth_entries": len(categorized['bluetooth'])
            },
            "case_information": {},
            "system_information": system_info,
            "extraction_information": extraction_info,
            "location_data": [],
            "speed_data": [],
            "bluetooth_data": []
        }
        
        # Add case information if provided
        if examiner_name:
            json_data["case_information"]["examiner_name"] = examiner_name
        if case_number:
            json_data["case_information"]["case_number"] = case_number
        
        # Process location data
        location_headers = self.get_location_headers()
        for entry in categorized['location']:
            row = self.format_location_entry_for_xlsx(entry)
            entry_dict = {}
            
            for i, header in enumerate(location_headers):
                if i < len(row):
                    entry_dict[header] = row[i]
            
            entry_dict.update({
                "raw_latitude": entry.latitude,
                "raw_longitude": entry.longitude,
                "raw_timestamp": entry.timestamp,
                "raw_extra_data": entry.extra_data
            })
            
            json_data["location_data"].append(entry_dict)
        
        # Process speed data
        speed_headers = self.get_speed_headers()
        for entry in categorized['speed']:
            row = self.format_speed_entry_for_xlsx(entry)
            entry_dict = {}
            
            for i, header in enumerate(speed_headers):
                if i < len(row):
                    entry_dict[header] = row[i]
            
            entry_dict.update({
                "raw_timestamp": entry.timestamp,
                "raw_extra_data": entry.extra_data
            })
            
            json_data["speed_data"].append(entry_dict)
          # Process bluetooth data
        bluetooth_headers = self.get_bluetooth_headers()
        for entry in categorized['bluetooth']:
            row = self.format_bluetooth_entry_for_xlsx(entry)
            entry_dict = {}
            
            for i, header in enumerate(bluetooth_headers):
                if i < len(row):
                    entry_dict[header] = row[i]
            
            entry_dict.update({
                "raw_timestamp": entry.timestamp,
                "raw_extra_data": entry.extra_data
            })
            
            json_data["bluetooth_data"].append(entry_dict)
        
        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False, default=str)
        
        self._logger.info(f"JSON report with separate sections written successfully: {output_path}")